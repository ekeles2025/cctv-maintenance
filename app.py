import os
import io
import csv
import time
import logging
import sqlite3
import re
import subprocess
from datetime import datetime, timedelta, timezone
from dotenv import load_dotenv
import pandas as pd
from werkzeug.utils import secure_filename

from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, session, g, make_response
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from flask_wtf.csrf import CSRFProtect
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.exceptions import HTTPException
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

try:
    from flask_babel import Babel, gettext as _
    from flask_babel import get_locale
    BABEL_AVAILABLE = True
except ImportError:  # pragma: no cover
    def _(x): return x
    BABEL_AVAILABLE = False
    get_locale = None

# Import configuration and utilities
from config import Config
from utils import utc_now, local_now, local_dt, allowed_upload_file, ensure_directories, calculate_time_diff, check_file_size
from menu_config import load_menu_items, save_menu_items, update_menu_item

# Set environment variables directly
os.environ['SECRET_KEY'] = 'dev-secret-key-change-in-production'

# إنشاء التطبيق
app = Flask(__name__)
app.config.from_object(Config)

# إعداد قاعدة بيانات PostgreSQL كأساسي دائماً - يتجاوز أي إعدادات في Config
database_url = os.environ.get('DATABASE_URL', 'postgresql://postgres:SOcgHFJDqcDrvUKzermleZkoMPbjmmxC@postgres-bvfp.railway.internal:5432/railway')

# التحقق من توفر psycopg2 لـ PostgreSQL
try:
    import psycopg2
    app.config["SQLALCHEMY_DATABASE_URI"] = database_url
    print("Using PostgreSQL database")
except ImportError:
    print("psycopg2 not found, falling back to SQLite")
    print("Please add 'psycopg2-binary>=2.9.0' to requirements.txt")
    app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///camera_system.db"

app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db = SQLAlchemy(app)

# تهيئة قاعدة البيانات تلقائياً عند بدء التشغيل
def auto_init_database():
    """Initialize database automatically on startup"""
    with app.app_context():
        try:
            # إنشاء الجداول إذا لم تكن موجودة
            db.create_all()
            
            # التحقق من وجود مستخدم admin
            if not User.query.filter_by(username='admin').first():
                admin_user = User(
                    username='admin',
                    password=generate_password_hash('Mostafa@2025', method='sha256'),
                    role='admin'
                )
                db.session.add(admin_user)
                db.session.commit()
                print("Admin user created successfully!")
            
        except Exception as e:
            print(f"Database initialization error: {e}")
            db.session.rollback()

# تشغيل التهيئة التلقائية
auto_init_database()

SUPPORTED_LANGS = {"ar", "en"}

TRANSLATIONS = {
    "ar": {
        "system_subtitle": "Camera System",
        "home": "الرئيسية",
        "faults": "الأعطال",
        "faulty_devices": "الأجهزة المعطلة",
        "import_excel": "استيراد من Excel",
        "chains": "السلاسل",
        "regions": "المناطق",
        "technicians": "الفنيين",
        "settings": "الإعدادات",
        "logout": "تسجيل الخروج",
        "admin": "مدير",
        "technician": "فني",
        "login_title": "تسجيل الدخول",
        "login_system": "نظام إدارة الكاميرات",
        "username": "اسم المستخدم",
        "password": "كلمة المرور",
        "sign_in": "تسجيل الدخول",
        "username_placeholder": "أدخل اسم المستخدم",
        "password_placeholder": "أدخل كلمة المرور",
        "ping_on_store": "Ping On Store",
        "test_connection": "اختبار الاتصال",
        "search_branch": "البحث عن الفرع",
        "enter_ip_address": "أدخل عنوان IP",
        "invalid_ip": "عنوان IP غير صحيح",
        "connection_success": "تم الاتصال بنجاح",
        "connection_failed": "فشل الاتصال",
        "connection_timeout": "انتهت مهلة الاتصال",
        "connection_error": "خطأ في الاتصال",
        "available_branches": "الفروع المتاحة",
        "no_matching_branches": "لا توجد فروع مطابقة للبحث",
        "try_different_search": "جرب البحث بكلمات مختلفة أو تحقق من إضافة الفروع",
        "select_device": "اختر الجهاز للاتصال",
        "devices_loaded_auto": "سيتم تحميل الأجهزة تلقائياً عند اختيار الفرع",
        "enter_ip_or_select": "أدخل عنوان IP مباشرة أو اختر جهازاً من القائمة",
        "ping_result": "نتيجة الاختبار",
        "ping_success": "نجح الاتصال!",
        "ping_failed": "فشل الاتصال",
        "start_test": "ابدأ الاختبار",
        "cancel": "إلغاء",
        "search_by_branch": "ابحث باسم الفرع، المنطقة، أو السلسلة",
        "example_branch_search": "مثال: فرع القاهرة، السلسلة الرئيسية...",
        "search": "بحث",
        "clear_search": "مسح البحث",
        "no_chain": "بدون سلسلة",
        "no_region": "بدون منطقة",
        "no_location": "لا يوجد موقع محدد",
        "available_devices": "الأجهزة المتاحة:",
        "no_devices_with_ip": "لا توجد أجهزة بعناوين IP",
        "response_time": "وقت الاستجابة",
        "connected": "متصل",
        "not_connected": "غير متصل",
        "unavailable": "غير متوفر",
        "ping_details": "تفاصيل الـ Ping",
        "connection_error": "خطأ في الاتصال",
        "status": "الحالة",
        "time": "الزمن",
        "testing": "جاري الاختبار",
        "very_fast": "سريع جداً",
        "good": "جيد",
        "slow": "بطيء",
        "very_slow": "بطيء جداً",
        "timeout": "انتهت المهلة",
        "open_cmd_confirm": "فتح CMD للتأكيد",
        "enter_valid_ip": "الرجاء إدخال عنوان IP صحيح",
        "copied_to_clipboard": "تم نسخ الأمر إلى الحافظة. افتح CMD والصق الأمر:",
        "copy_command": "انسخ هذا الأمر وافتح CMD:",
        "cmd_results": "نتائج CMD",
        "ping_10_seconds": "جاري تشغيل اختبار Ping لمدة 10 ثواني...",
        "ping_completed": "اكتمل اختبار Ping. جاري تحميل النتائج...",
        "ping_test_completed": "تم تشغيل اختبار Ping لمدة 10 ثواني على",
        "copy_results": "نسخ النتائج",
        "download_results": "تحميل النتائج",
        "close": "إغلاق",
        "results_copied": "تم نسخ النتائج إلى الحافظة",
        "powershell_script_created": "PowerShell script created. Run it to perform the ping test.",
        "cmd_open_failed": "Could not open CMD automatically. Please copy the command and run it manually.",
        "cmd_open_error": "Error opening CMD. Command copied to clipboard.",
        "camera_offline": "Camera Offline",
        "other": "أخرى",
        "سيتم حفظ الملاحظات في حقل الملاحظات الخاص بالكاميرا": "سيتم حفظ الملاحظات في حقل الملاحظات الخاص بالكاميرا",
        "default_accounts": "الحسابات الافتراضية",
        "login_information": "معلومات الدخول",
        "admin_colon": "مدير:",
        "technician_colon": "فني:",
        "signing_in": "جاري تسجيل الدخول...",
        "قائمة أعطال الكاميرات": "قائمة أعطال الكاميرات",
        "ابحث عن عطل...": "ابحث عن عطل...",
        "حالة العطل": "حالة العطل",
        "الكل": "الكل",
        "مفتوحة": "مفتوحة",
        "مُصلحة": "مُصلحة",
        "الفرع": "الفرع",
        "كل الفروع": "كل الفروع",
        "إحصائيات سريعة": "إحصائيات سريعة",
        "الإجمالي:": "الإجمالي:",
        "مفتوحة:": "مفتوحة:",
        "مُصلحة:": "مُصلحة:",
        "إدارة أعطال الكاميرات": "إدارة أعطال الكاميرات",
        "تحديث": "تحديث",
        "عطل جديد": "عطل جديد",
        "الكاميرا": "الكاميرا",
        "نوع العطل": "نوع العطل",
        "الوصف": "الوصف",
        "الحالة": "الحالة",
        "التاريخ": "التاريخ",
        "الإجراءات": "الإجراءات",
        "جاري التحميل...": "جاري التحميل...",
        "لا توجد أعطال": "لا توجد أعطال",
        "لم يتم العثور على أي أعطال مطابقة للبحث": "لم يتم العثور على أي أعطال مطابقة للبحث",
        "تعديل العطل": "تعديل العطل",
        "اختر نوع العطل": "اختر نوع العطل",
        "كهرباء": "كهرباء",
        "شبكة": "شبكة",
        "جهاز": "جهاز",
        "برمجيات": "برمجيات",
        "كابل": "كابل",
        "محروق": "محروق",
        "نوع الجهاز": "نوع الجهاز",
        "اختر نوع الجهاز": "اختر نوع الجهاز",
        "كاميرا": "كاميرا",
        "المبلغ": "المبلغ",
        "الفني المسؤول": "الفني المسؤول",
        "لم يتم التعيين": "لم يتم التعيين",
        "ملاحظات الإصلاح": "ملاحظات الإصلاح",
        "إلغاء": "إلغاء",
        "حفظ التعديلات": "حفظ التعديلات",
        "تأكيد إصلاح العطل": "تأكيد إصلاح العطل",
        "هل أنت متأكد من إصلاح هذا العطل؟": "هل أنت متأكد من إصلاح هذا العطل؟",
        "اكتب هنا تفاصيل الإصلاح...": "اكتب هنا تفاصيل الإصلاح...",
        "صورة الإصلاح (اختياري)": "صورة الإصلاح (اختياري)",
        "تأكيد الإصلاح": "تأكيد الإصلاح",
        "حدث خطأ في تحميل الأعطال": "حدث خطأ في تحميل الأعطال",
        "محول الي BPM": "محول الي BPM",
        "مُصلح": "مُصلح",
        "قيد الانتظار": "قيد الانتظار",
        "تعديل": "تعديل",
        "تحويل إلى BPM": "تحويل إلى BPM",
        "إصلاح": "إصلاح",
        "عرض": "عرض",
        "حذف": "حذف",
        "تم تعديل العطل بنجاح": "تم تعديل العطل بنجاح",
        "حدث خطأ": "حدث خطأ",
        "حدث خطأ في حفظ التعديلات": "حدث خطأ في حفظ التعديلات",
        "تم إصلاح العطل بنجاح": "تم إصلاح العطل بنجاح",
        "حدث خطأ في إصلاح العطل": "حدث خطأ في إصلاح العطل",
        "هل أنت متأكد من تحويل هذا العطل إلى BPM؟ سيتم نقله إلى قسم BPM ولن يظهر في قائمة الأعطال الرئيسية": "هل أنت متأكد من تحويل هذا العطل إلى BPM؟ سيتم نقله إلى قسم BPM ولن يظهر في قائمة الأعطال الرئيسية",
        "تم تحويل العطل إلى BPM بنجاح": "تم تحويل العطل إلى BPM بنجاح",
        "حدث خطأ في تحويل العطل إلى BPM": "حدث خطأ في تحويل العطل إلى BPM",
        "هل أنت متأكد من حذف هذا العطل؟ لا يمكن التراجع عن هذا الإجراء": "هل أنت متأكد من حذف هذا العطل؟ لا يمكن التراجع عن هذا الإجراء",
        "تم حذف العطل بنجاح": "تم حذف العطل بنجاح",
        "حدث خطأ في حذف العطل": "حدث خطأ في حذف العطل",
        "مغلق": "مغلق",
        "مفتوح": "مفتوح",
        "في": "في",
        "تقارير فتح وإغلاق الفروع": "تقارير فتح وإغلاق الفروع",
        "تنزيل تقارير بيانات فتح وإغلاق الفروع": "تنزيل تقارير بيانات فتح وإغلاق الفروع",
        "تقرير فتح وإغلاق الفروع": "تقرير فتح وإغلاق الفروع",
    },
    "en": {
        "system_subtitle": "Camera System",
        "home": "Home",
        "faults": "Faults",
        "faulty_devices": "Faulty Devices",
        "import_excel": "Import Excel",
        "chains": "Chains",
        "regions": "Regions",
        "technicians": "Technicians",
        "settings": "Settings",
        "logout": "Logout",
        "admin": "Admin",
        "technician": "Technician",
        "login_title": "Sign In",
        "login_system": "Camera Management System",
        "username": "Username",
        "password": "Password",
        "sign_in": "Sign In",
        "username_placeholder": "Enter username",
        "password_placeholder": "Enter password",
        "ping_on_store": "Ping On Store",
        "test_connection": "Test Connection",
        "search_branch": "Search Branch",
        "enter_ip_address": "Enter IP Address",
        "invalid_ip": "Invalid IP Address",
        "connection_success": "Connection Successful",
        "connection_failed": "Connection Failed",
        "connection_timeout": "Connection Timeout",
        "connection_error": "Connection Error",
        "available_branches": "Available Branches",
        "no_matching_branches": "No branches match your search",
        "try_different_search": "Try searching with different keywords or check if branches are added",
        "select_device": "Select Device to Connect",
        "devices_loaded_auto": "Devices will load automatically when you select a branch",
        "enter_ip_or_select": "Enter IP address directly or select a device from the list",
        "ping_result": "Test Result",
        "ping_success": "Connection Successful!",
        "ping_failed": "Connection Failed",
        "start_test": "Start Test",
        "cancel": "Cancel",
        "search_by_branch": "Search by branch name, region, or chain",
        "example_branch_search": "Example: Cairo Branch, Main Chain...",
        "search": "Search",
        "clear_search": "Clear Search",
        "no_chain": "No Chain",
        "no_region": "No Region",
        "no_location": "No location specified",
        "available_devices": "Available Devices:",
        "no_devices_with_ip": "No devices with IP addresses",
        "response_time": "Response Time",
        "connected": "Connected",
        "not_connected": "Not Connected",
        "unavailable": "Unavailable",
        "ping_details": "Ping Details",
        "connection_error": "Connection Error",
        "status": "Status",
        "time": "Time",
        "testing": "Testing",
        "very_fast": "Very Fast",
        "good": "Good",
        "slow": "Slow",
        "very_slow": "Very Slow",
        "timeout": "Timeout",
        "open_cmd_confirm": "Open CMD for Confirmation",
        "enter_valid_ip": "Please enter a valid IP address",
        "copied_to_clipboard": "Command copied to clipboard. Open CMD and paste:",
        "copy_command": "Copy this command and open CMD:",
        "cmd_results": "CMD Results",
        "ping_10_seconds": "Running ping test for 10 seconds...",
        "ping_completed": "Ping test completed. Loading results...",
        "ping_test_completed": "Ping test completed for 10 seconds on",
        "copy_results": "Copy Results",
        "download_results": "Download Results",
        "close": "Close",
        "results_copied": "Results copied to clipboard",
        "powershell_script_created": "PowerShell script created. Run it to perform the ping test.",
        "cmd_open_failed": "Could not open CMD automatically. Please copy the command and run it manually.",
        "cmd_open_error": "Error opening CMD. Command copied to clipboard.",
        "camera_offline": "Camera Offline",
        "other": "Other",
        "default_accounts": "Default Accounts",
        "login_information": "Login Information",
        "admin_colon": "Admin:",
        "technician_colon": "Technician:",
        "signing_in": "Signing in...",
        "login_page_title": "Login Page",
        "login_page_subtitle": "Enter your credentials to access the system",
        "username_label": "Username",
        "password_label": "Password",
        "remember_me": "Remember Me",
        "login_button": "Login",
        "forgot_password": "Forgot Password",
        "register_now": "Register Now",
        "faults_list": "Faults List",
        "search_fault": "Search for a fault...",
        "device_management": "Device Management",
        "device_faults": "Device Faults",
        "ping_on_store": "Ping On Store",
        "import_excel": "Import Excel",
        "reports": "Reports",
        "status": "Status",
        "date": "Date",
        "actions": "Actions",
        "loading": "Loading...",
        "no_faults": "No Faults",
        "no_faults_found": "No faults found matching the search",
        "edit_fault": "Edit Fault",
        "select_fault_type": "Select Fault Type",
        "electricity": "Electricity",
        "network": "Network",
        "device": "Device",
        "software": "Software",
        "cable": "Cable",
        "burned": "Burned",
        "other": "Other",
        "device_type": "Device Type",
        "select_device_type": "Select Device Type",
        "camera": "Camera",
        "reporter": "Reporter",
        "all_branches": "All Branches",
        "quick_statistics": "Quick Statistics",
        "total": "Total:",
        "open": "Open:",
        "resolved": "Resolved:",
        "fault_management": "Fault Management",
        "refresh": "Refresh",
        "new_fault": "New Fault",
        "camera": "Camera",
        "fault_type": "Fault Type",
        "description": "Description",
        "status": "Status",
        "date": "Date",
        "actions": "Actions",
        "loading": "Loading...",
        "no_faults": "No Faults",
        "no_faults_found": "No faults found matching the search",
        "edit_fault": "Edit Fault",
        "select_fault_type": "Select Fault Type",
        "electricity": "Electricity",
        "not_assigned": "Not Assigned",
        "repair_notes": "Repair Notes",
        "cancel": "Cancel",
        "save_changes": "Save Changes",
        "confirm_repair": "Confirm Repair",
        "confirm_repair_question": "Are you sure you want to mark this fault as repaired?",
        "write_repair_details": "Write repair details here...",
        "repair_image_optional": "Repair Image (Optional)",
        "confirm_repair_button": "Confirm Repair",
        "error_loading_faults": "Error loading faults",
        "converted_to_bbm": "Converted to BPM",
        "resolved": "Resolved",
        "waiting": "Waiting",
        "edit": "Edit",
        "convert_to_bbm": "Convert to BPM",
        "repair": "Repair",
        "view": "View",
        "delete": "Delete",
        "fault_updated_success": "Fault updated successfully",
        "error_occurred": "An error occurred",
        "error_saving_changes": "Error saving changes",
        "fault_repaired_success": "Fault repaired successfully",
        "error_repairing_fault": "Error repairing fault",
        "confirm_convert_to_bbm": "Are you sure you want to convert this fault to BPM? It will be moved to the BPM department and will not appear in the main faults list",
        "fault_converted_to_bbm_success": "Fault successfully converted to BPM",
        "error_converting_to_bbm": "Error converting fault to BPM",
        "confirm_delete_fault": "Are you sure you want to delete this fault? This action cannot be undone.",
        "fault_deleted_success": "Fault deleted successfully",
        "error_deleting_fault": "Error deleting fault",
        "branch_phone_numbers_data": "Branch Phone Numbers Data",
        "error_uploading_phone_numbers": "Error uploading phone numbers",
        "Branch Status": "Branch Status",
        "Open": "Open",
        "Closed": "Closed",
        "On": "On",
        "Branch Open/Close Reports": "Branch Open/Close Reports",
        "Download branch open/close data reports": "Download branch open/close data reports",
        "Branch Open/Close Report": "Branch Open/Close Report",
    },
}


def get_current_lang():
    # Force English as default language
    if hasattr(request, 'session'):
        lang = request.session.get("lang", "en")
    else:
        # Fallback to direct session access
        from flask import session
        lang = session.get("lang", "en")
    
    # Always default to English if not explicitly set to Arabic
    if lang != "ar":
        lang = "en"
    
    return lang if lang in SUPPORTED_LANGS else "en"


babel = Babel(app, locale_selector=get_current_lang) if BABEL_AVAILABLE else None

try:
    app.jinja_env.add_extension('jinja2.ext.i18n')
    # Install gettext function in Jinja2 environment
    if BABEL_AVAILABLE:
        from flask_babel import gettext
        app.jinja_env.install_gettext_callables(
            gettext,  # gettext function
            lambda s, p, n: ngettext(s, p, n),  # ngettext function
            newstyle=True
        )
except Exception:
    pass


@app.before_request
def set_language_globals():
    g.lang = get_current_lang()
    g.dir = "rtl" if g.lang == "ar" else "ltr"
    
    # Force English as default for new sessions
    if 'lang' not in session:
        session['lang'] = 'en'
        g.lang = 'en'
        g.dir = 'ltr'


@app.route("/")
def home():
    return redirect(url_for("dashboard"))

@app.route("/health")
def health_check():
    """Health check endpoint for Docker"""
    return jsonify({"status": "healthy", "timestamp": datetime.now().isoformat()})

@app.route("/review")
def review_redirect():
    """Redirect /review to /settings to fix 404 errors"""
    return redirect(url_for("settings"))

@app.route("/set-language/<lang>")
def set_language(lang):
    if lang not in SUPPORTED_LANGS:
        lang = "en"
    session["lang"] = lang

    next_url = request.args.get("next")
    if next_url and next_url.startswith("/"):
        return redirect(next_url)
    return redirect(request.referrer or url_for("dashboard"))


def t(key: str) -> str:
    lang = getattr(g, "lang", "en")
    return TRANSLATIONS.get(lang, {}).get(key, TRANSLATIONS["en"].get(key, key))

# حماية CSRF
csrf = CSRFProtect(app)

# إعداد Logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logging.Formatter.converter = time.localtime
logger = logging.getLogger(__name__)

# Create directories and initialize paths
ensure_directories(app.config['UPLOAD_FOLDER'], app.config['LOGO_FOLDER'])

# Import menu model after db is initialized
from menu_model import MenuItem
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"

# تعريف النماذج
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(100), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)
    role = db.Column(db.String(20), default="technician")
    faults_reported = db.relationship('Fault', backref='technician', lazy=True)
    device_faults_reported = db.relationship('DeviceFault', backref='technician', lazy=True)

    def get_id(self):
        return str(self.id)

    @property
    def is_authenticated(self):
        return True

    @property
    def is_active(self):
        return True

    @property
    def is_anonymous(self):
        return False

def _looks_like_ip(text):
    """Check if text looks like an IP address"""
    if not text:
        return False
    
    # Split by common delimiters
    parts = text.replace('.', ' ').replace('-', ' ').split()
    
    # Check if we have 4 parts (typical IP format)
    if len(parts) == 4:
        # Check if all parts are numbers or valid IP parts
        for part in parts:
            if not part.isdigit() and part != '0':  # Allow 0 but check other parts
                return False
        return True
    
    return False

class Chain(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    regions = db.relationship("Region", backref="chain", lazy=True, cascade="all, delete-orphan")

class Region(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    chain_id = db.Column(db.Integer, db.ForeignKey('chain.id'), nullable=True)
    branches = db.relationship("Branch", backref="region", cascade="all, delete")

class Branch(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    location = db.Column(db.String(200))
    ip_address = db.Column(db.String(100))
    phone_number = db.Column(db.String(20))
    phone_number_2 = db.Column(db.String(20))
    phone_number_3 = db.Column(db.String(20))
    phone_number_4 = db.Column(db.String(20))
    nvr_count = db.Column(db.Integer, default=0)
    branch_type = db.Column(db.String(20), default='دائم')  # دائم أو موسمي
    sequence_number = db.Column(db.Integer, default=0)  # Auto sequence number
    region_id = db.Column(db.Integer, db.ForeignKey('region.id'))
    cameras = db.relationship("Camera", backref="branch", cascade="all, delete")
    devices = db.relationship("Device", backref="branch", cascade="all, delete")
    
    # Branch closure fields
    closed = db.Column(db.Boolean, default=False)
    closure_reason = db.Column(db.Text)
    reporter_name = db.Column(db.Text)
    closure_date = db.Column(db.DateTime(timezone=True))
    
    # Relationship with history
    history = db.relationship("BranchHistory", backref="branch", cascade="all, delete")

class BranchHistory(db.Model):
    """Track branch open/close history"""
    id = db.Column(db.Integer, primary_key=True)
    branch_id = db.Column(db.Integer, db.ForeignKey('branch.id'), nullable=False)
    action = db.Column(db.String(10), nullable=False)  # 'open' or 'close'
    action_date = db.Column(db.DateTime(timezone=True), nullable=False, default=utc_now)
    reason = db.Column(db.Text)
    reporter_name = db.Column(db.Text)
    created_at = db.Column(db.DateTime(timezone=True), default=utc_now)

class ExcelData(db.Model):
    """Model to store raw Excel data for Total Camera"""
    id = db.Column(db.Integer, primary_key=True)
    excel_filename = db.Column(db.String(255), nullable=False)
    upload_date = db.Column(db.DateTime(timezone=True), default=utc_now)
    row_data = db.Column(db.Text, nullable=False)  # JSON string of row data
    row_number = db.Column(db.Integer, nullable=False)

    def to_dict(self):
        """Convert row data back to dictionary"""
        try:
            import json
            return json.loads(self.row_data)
        except:
            return {}

class Camera(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    ip_address = db.Column(db.String(100))
    camera_type = db.Column(db.String(50), default="مراقبة")
    status = db.Column(db.String(20), default="online")  # online, offline, maintenance
    branch_id = db.Column(db.Integer, db.ForeignKey('branch.id'))
    sequence_number = db.Column(db.Integer, default=0)  # Auto sequence number
    faults = db.relationship("Fault", backref="camera", cascade="all, delete")
    
    # Additional fields for Total Camera report
    ops_area = db.Column(db.String(100))  # Operations Area
    total_cameras = db.Column(db.Integer, default=0)  # Total cameras count
    serious_cameras = db.Column(db.Integer, default=0)  # Serious cameras count
    indoor_cameras = db.Column(db.Integer, default=0)  # Indoor cameras count
    outdoor_cameras = db.Column(db.Integer, default=0)  # Outdoor cameras count
    nvr_count = db.Column(db.Integer, default=0)  # NVR count
    note = db.Column(db.String(500))  # Additional notes

class Fault(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    description = db.Column(db.String(200), nullable=False)
    fault_type = db.Column(db.String(100), nullable=False)
    device_type = db.Column(db.String(50), nullable=False, default="NVR")  # NVR, Switch, Router, Up Link
    reported_by = db.Column(db.String(100), nullable=False)
    technician_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    date_reported = db.Column(db.DateTime(timezone=True), default=utc_now)
    expires_at = db.Column(db.DateTime(timezone=True), default=lambda: utc_now() + timedelta(days=7))
    resolved_at = db.Column(db.DateTime(timezone=True), nullable=True)
    resolved_by = db.Column(db.String(100), nullable=True)
    repair_notes = db.Column(db.String(300), nullable=True)
    repair_image = db.Column(db.String(200), nullable=True)
    camera_id = db.Column(db.Integer, db.ForeignKey('camera.id'))
    
    # الحقول الجديدة
    camera_name = db.Column(db.String(100))
    problem = db.Column(db.String(200))

class Device(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    device_type = db.Column(db.String(50), nullable=False)  # NVR, Switch, Router, Up Link
    ip_address = db.Column(db.String(100))
    location = db.Column(db.String(200))
    branch_id = db.Column(db.Integer, db.ForeignKey('branch.id'))
    faults = db.relationship("DeviceFault", backref="device", cascade="all, delete")

class DeviceFault(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    description = db.Column(db.String(200), nullable=False)
    fault_type = db.Column(db.String(100), nullable=False)
    reported_by = db.Column(db.String(100), nullable=False)
    technician_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    date_reported = db.Column(db.DateTime(timezone=True), default=utc_now)
    expires_at = db.Column(db.DateTime(timezone=True), default=lambda: utc_now() + timedelta(days=7))
    resolved_at = db.Column(db.DateTime(timezone=True), nullable=True)
    resolved_by = db.Column(db.String(100), nullable=True)
    repair_notes = db.Column(db.String(300), nullable=True)
    repair_image = db.Column(db.String(200), nullable=True)
    device_id = db.Column(db.Integer, db.ForeignKey('device.id'))

class BBMFault(db.Model):
    """Model to store faults transferred to BBM"""
    id = db.Column(db.Integer, primary_key=True)
    original_fault_id = db.Column(db.Integer, nullable=False)  # Reference to original fault
    camera_id = db.Column(db.Integer, db.ForeignKey('camera.id'), nullable=False)
    branch_id = db.Column(db.Integer, db.ForeignKey('branch.id'), nullable=False)
    description = db.Column(db.String(200), nullable=False)
    fault_type = db.Column(db.String(100), nullable=False)
    device_type = db.Column(db.String(50), nullable=False, default="NVR")
    reported_by = db.Column(db.String(100), nullable=False)
    technician_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    date_reported = db.Column(db.DateTime(timezone=True), default=utc_now)
    date_transferred = db.Column(db.DateTime(timezone=True), default=utc_now)
    transferred_by = db.Column(db.String(100), nullable=False)
    status = db.Column(db.String(50), nullable=False, default="Pending")  # Pending, In Progress, Resolved
    notes = db.Column(db.String(500), nullable=True)
    
    # Relationships
    camera = db.relationship('Camera', backref='bbm_faults')
    branch = db.relationship('Branch', backref='bbm_faults')
    technician = db.relationship('User', backref='bbm_faults')

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


def ensure_device_type_column():
    """Ensure `device_type` column exists in `fault` table for SQLite databases.
    This is a safe one-time operation: adds the column and fills existing rows with 'NVR'.
    """
    try:
        uri = app.config.get('SQLALCHEMY_DATABASE_URI', '')
        if not uri.startswith('sqlite'):
            return

        # Support sqlite:///relative.db or sqlite:////absolute/path.db
        if uri.startswith('sqlite:///'):
            db_path = uri.replace('sqlite:///', '')
        else:
            db_path = uri.replace('sqlite://', '')

        db_path = os.path.abspath(db_path)
        if not os.path.exists(db_path):
            logger.info(f"SQLite DB not found at {db_path}; skipping device_type check")
            return

        conn = sqlite3.connect(db_path)
        cur = conn.cursor()
        cur.execute("PRAGMA table_info(fault)")
        cols = [r[1] for r in cur.fetchall()]
        if 'device_type' in cols:
            logger.info('device_type column already present in fault table')
            conn.close()
            return

        logger.info('Adding device_type column to fault table (one-time)')
        cur.execute("ALTER TABLE fault ADD COLUMN device_type TEXT;")
        cur.execute("UPDATE fault SET device_type = 'NVR' WHERE device_type IS NULL OR device_type = '';")
        conn.commit()
        conn.close()
        logger.info('device_type column added and populated with default NVR')

    except Exception as exc:
        logger.error(f"Failed ensuring device_type column: {exc}")


# Configure Babel
app.config['BABEL_DEFAULT_LOCALE'] = 'en'
app.config['BABEL_DEFAULT_TIMEZONE'] = 'UTC'
app.config['BABEL_TRANSLATION_DIRECTORIES'] = 'translations'
app.config['TEMPLATES_AUTO_RELOAD'] = True
app.config['WTF_CSRF_ENABLED'] = True

@app.context_processor
def inject_logo():
    logo_exists = os.path.exists(app.config['COMPANY_LOGO'])
    company_name = app.config.get('COMPANY_NAME', 'CCTV Portal EG')
    logo_size = app.config.get('LOGO_SIZE', 50)
    
    # Force English as default
    current_lang = getattr(g, 'lang', 'en')
    if current_lang != 'en' and current_lang != 'ar':
        current_lang = 'en'
    
    current_dir = 'ltr' if current_lang == 'en' else 'rtl'
    
    return dict(
        logo_exists=logo_exists,
        company_name=company_name,
        logo_size=logo_size,
        current_lang=current_lang,
        current_dir=current_dir,
        t=t,
        _=_ ,
        time=time,
    )


# Register Jinja filter for local datetime formatting
app.jinja_env.filters['local_dt'] = local_dt

# Register Jinja filter to remove .0 from decimal numbers
def remove_decimal(value):
    """Remove .0 from decimal numbers, return original value if not a number"""
    if value is None:
        return value
    try:
        # Convert to float and check if it's a whole number
        if isinstance(value, (int, float)):
            if float(value) == int(float(value)):
                return str(int(float(value)))
        # If it's a string, try to convert to number
        if isinstance(value, str):
            try:
                num = float(value)
                if num == int(num):
                    return str(int(num))
                return value
            except ValueError:
                return value
        return value
    except (ValueError, TypeError):
        return value

app.jinja_env.filters['remove_decimal'] = remove_decimal

# Error handlers
@app.errorhandler(404)
def not_found_error(error):
    logger.warning(f"404 error: {request.path}")
    flash(_("الصفحة المطلوبة غير موجودة"), "danger")
    return redirect(url_for("dashboard"))

@app.errorhandler(500)
def internal_error(error):
    logger.error(f"500 error: {str(error)}", exc_info=True)
    flash("A server error occurred. Please try again later", "danger")
    return redirect(url_for("dashboard"))

@app.errorhandler(403)
def forbidden_error(error):
    logger.warning(f"403 error: {request.path}")
    flash(_("لا تملك صلاحية للوصول لهذه الصفحة"), "danger")
    return redirect(url_for("dashboard"))

# الروتات
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form['username']
        password = request.form['password']
        user = User.query.filter_by(username=username).first()
        if user and check_password_hash(user.password, password):
            login_user(user)
            flash(_("تم تسجيل الدخول بنجاح"), "success")
            return redirect(url_for("dashboard"))
        else:
            flash(_("اسم المستخدم أو كلمة المرور غير صحيحة"), "danger")
    return render_template("login.html")

@app.route("/logout")
@login_required
def logout():
    logout_user()
    flash(_("تم تسجيل الخروج بنجاح"), "info")
    return redirect(url_for("login"))
@app.route("/dashboard")
@login_required
def dashboard():
    now = local_now()
    total_cameras = Camera.query.count()  # Count all cameras, not just those in branches
    total_devices = Device.query.count()
    
    # Count actual fault records
    open_faults = Fault.query.filter(Fault.resolved_at == None).count()
    closed_faults = Fault.query.filter(Fault.resolved_at != None).count()
    open_device_faults = DeviceFault.query.filter(DeviceFault.resolved_at == None).count()
    closed_device_faults = DeviceFault.query.filter(DeviceFault.resolved_at != None).count()
    
    # Count branches
    total_branches = Branch.query.count()
    closed_branches = Branch.query.filter(Branch.closed == True).count()
    open_branches = total_branches - closed_branches
    
    # Count unique cameras/devices with faults (from actual fault records)
    faulty_cameras = db.session.query(Fault.camera_id).filter(
        Fault.resolved_at == None
    ).distinct().count()
    faulty_devices = db.session.query(DeviceFault.device_id).filter(
        DeviceFault.resolved_at == None
    ).distinct().count()
    
    # Debug logging
    logger.info(f"Dashboard stats - Total cameras: {total_cameras}, Total devices: {total_devices}")
    logger.info(f"Faults - Open: {open_faults}, Closed: {closed_faults}, Faulty cameras: {faulty_cameras}, Faulty devices: {faulty_devices}")
    logger.info(f"Device faults - Open: {open_device_faults}, Closed: {closed_device_faults}")
    
    working_cameras = total_cameras - faulty_cameras

    open_faults_list = Fault.query.filter(
        Fault.resolved_at == None
    ).order_by(Fault.date_reported.asc()).limit(10).all()  # Get oldest 10 open faults
    
    # Calculate durations for recent faults
    recent_faults_with_duration = []
    for fault in open_faults_list:
        duration = calculate_time_diff(fault.date_reported, now)
        recent_faults_with_duration.append({
            'fault': fault,
            'duration': duration
        })
    
    open_device_faults_list = DeviceFault.query.filter(
        DeviceFault.resolved_at == None
    ).order_by(DeviceFault.date_reported.asc()).limit(10).all()

    my_faults = []
    my_device_faults = []
    if current_user.role == "technician":
        my_faults = Fault.query.filter(
            Fault.technician_id == current_user.id,
            Fault.resolved_at == None
        ).order_by(Fault.date_reported.asc()).all()
        
        my_device_faults = DeviceFault.query.filter(
            DeviceFault.technician_id == current_user.id,
            DeviceFault.resolved_at == None
        ).order_by(DeviceFault.date_reported.asc()).all()

    return render_template("dashboard.html",
                           total_cameras=total_cameras,
                           total_devices=total_devices,
                           faulty_cameras=faulty_cameras,
                           working_cameras=working_cameras,
                           faulty_devices=faulty_devices,
                           open_faults=open_faults,
                           closed_faults=closed_faults,
                           open_device_faults=open_device_faults,
                           closed_device_faults=closed_device_faults,
                           total_branches=total_branches,
                           closed_branches=closed_branches,
                           open_branches=open_branches,
                           recent_faults_with_duration=recent_faults_with_duration,
                           recent_device_faults=open_device_faults_list,
                           my_faults=my_faults,
                           my_device_faults=my_device_faults,
                           now=now)

@app.route("/chains")
@login_required
def chains():
    if current_user.role != "admin":
        flash(_("فقط المدير يمكنه الوصول لهذه الصفحة"), "danger")
        return redirect(url_for("dashboard"))
    chains_list = Chain.query.all()
    chains_with_numbers = []
    for index, chain in enumerate(chains_list, 1):
        chains_with_numbers.append({
            'chain': chain,
            'row_number': index
        })
    return render_template("chains.html", chains=chains_with_numbers)

@app.route("/chains/add", methods=["GET", "POST"])
@login_required
def add_chain():
    if current_user.role != "admin":
        flash(_("فقط المدير يمكنه إضافة سلاسل"), "danger")
        return redirect(url_for("dashboard"))
    if request.method == "POST":
        name = request.form['name']
        chain = Chain(name=name)
        db.session.add(chain)
        db.session.commit()
        flash(_("تم إضافة السلسلة بنجاح"), "success")
        return redirect(url_for("chains"))
    return render_template("add_edit_chain.html", chain=None)

@app.route("/chains/edit/<int:chain_id>", methods=["GET", "POST"])
@login_required
def edit_chain(chain_id):
    if current_user.role != "admin":
        flash(_("فقط المدير يمكنه تعديل السلاسل"), "danger")
        return redirect(url_for("dashboard"))
    chain = Chain.query.get_or_404(chain_id)
    if request.method == "POST":
        chain.name = request.form['name']
        db.session.commit()
        flash(_("تم تعديل السلسلة بنجاح"), "success")
        return redirect(url_for("chains"))
    return render_template("add_edit_chain.html", chain=chain)

@app.route("/chains/delete/<int:chain_id>", methods=["POST"])
@login_required
def delete_chain(chain_id):
    if current_user.role != "admin":
        flash(_("فقط المدير يمكنه حذف السلاسل ❌"), "danger")
        return redirect(url_for("chains"))
    chain = Chain.query.get_or_404(chain_id)
    db.session.delete(chain)
    db.session.commit()
    flash(_("تم حذف السلسلة بنجاح"), "success")
    return redirect(url_for("chains"))

@app.route("/chains/<int:chain_id>/regions")
@login_required
def chain_regions(chain_id):
    if current_user.role != "admin":
        flash(_("فقط المدير يمكنه الوصول لهذه الصفحة ❌"), "danger")
        return redirect(url_for("dashboard"))
    chain = Chain.query.get_or_404(chain_id)
    regions_list = Region.query.filter_by(chain_id=chain_id).all()
    
    # Add sequential numbering
    regions_with_numbers = []
    for index, region in enumerate(regions_list, 1):
        regions_with_numbers.append({
            'region': region,
            'row_number': index
        })
    
    return render_template("regions.html", regions=regions_with_numbers, chain=chain)

@app.route("/chains/<int:chain_id>/regions/delete-all", methods=["POST"])
@login_required
def delete_all_chain_regions(chain_id):
    """حذف جميع المناطق في سلسلة معينة"""
    if current_user.role != "admin":
        flash(_("فقط المدير يمكنه حذف المناطق ❌"), "danger")
        return redirect(url_for("dashboard"))
    
    chain = Chain.query.get_or_404(chain_id)
    regions_count = Region.query.filter_by(chain_id=chain_id).count()
    
    if regions_count == 0:
        flash("لا توجد مناطق لحذفها في هذه السلسلة ℹ️", "info")
        return redirect(url_for("chain_regions", chain_id=chain_id))
    
    try:
        # حذف جميع المناطق (سيتم حذف الفروع والكاميرات والأجهزة تلقائياً بسبب cascade)
        Region.query.filter_by(chain_id=chain_id).delete()
        db.session.commit()
        
        flash(f"✅ Successfully deleted {regions_count} regions from chain '{chain.name}' with all related data (branches, cameras, devices, and faults)", "success")
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"خطأ في حذف مناطق السلسلة: {str(e)}")
        flash(f"حدث خطأ أثناء حذف المناطق: {str(e)} ❌", "danger")
    
    return redirect(url_for("chain_regions", chain_id=chain_id))

@app.route("/regions")
@login_required
def regions():
    if current_user.role != "admin":
        flash(_("فقط المدير يمكنه الوصول لهذه الصفحة ❌"), "danger")
        return redirect(url_for("dashboard"))
    chain_id = request.args.get("chain_id", type=int)
    if chain_id:
        regions_list = Region.query.filter_by(chain_id=chain_id).all()
        chain = Chain.query.get(chain_id)
    else:
        regions_list = Region.query.all()
        chain = None
    
    # Add sequential numbering
    regions_with_numbers = []
    for index, region in enumerate(regions_list, 1):
        regions_with_numbers.append({
            'region': region,
            'row_number': index
        })
    
    return render_template("regions.html", regions=regions_with_numbers, chain=chain)

@app.route("/regions/add", methods=["GET", "POST"])
@app.route("/regions/add/<int:chain_id>", methods=["GET", "POST"])
@login_required
def add_region(chain_id=None):
    if current_user.role != "admin":
        flash(_("فقط المدير يمكنه إضافة مناطق ❌"), "danger")
        return redirect(url_for("dashboard"))
    if request.method == "POST":
        name = request.form['name']
        cid = request.form.get('chain_id', type=int)
        region = Region(name=name, chain_id=cid)
        db.session.add(region)
        db.session.commit()
        flash("تم إضافة المنطقة بنجاح ✅", "success")
        if cid:
            return redirect(url_for("chain_regions", chain_id=cid))
        return redirect(url_for("regions"))
    chains_list = Chain.query.order_by(Chain.name).all()
    return render_template("add_edit_region.html", region=None, chains=chains_list, preselected_chain_id=chain_id)

@app.route("/regions/edit/<int:region_id>", methods=["GET", "POST"])
@login_required
def edit_region(region_id):
    if current_user.role != "admin":
        flash(_("فقط المدير يمكنه تعديل المناطق ❌"), "danger")
        return redirect(url_for("dashboard"))
    region = Region.query.get_or_404(region_id)
    if request.method == "POST":
        region.name = request.form['name']
        region.chain_id = request.form.get('chain_id', type=int) or None
        db.session.commit()
        flash("تم تعديل المنطقة بنجاح ✅", "success")
        if region.chain_id:
            return redirect(url_for("chain_regions", chain_id=region.chain_id))
        return redirect(url_for("regions"))
    chains_list = Chain.query.order_by(Chain.name).all()
    return render_template("add_edit_region.html", region=region, chains=chains_list, preselected_chain_id=None)

@app.route("/regions/delete-multiple", methods=["POST"])
@login_required
def delete_multiple_regions():
    """حذف المناطق المحددة دفعة واحدة"""
    if current_user.role != "admin":
        flash(_("غير مصرح لك بالوصول إلى هذه الصفحة"), "danger")
        return redirect(url_for("index"))
    
    region_ids = request.form.getlist("region_ids")
    
    if not region_ids:
        flash(_("الرجاء تحديد منطقة واحدة على الأقل"), "warning")
        return redirect(url_for("regions"))
    
    try:
        deleted_count = 0
        errors = []
        
        for region_id in region_ids:
            try:
                region = Region.query.get(region_id)
                if region:
                    # Get region name for message
                    region_name = region.name
                    
                    # Delete all related data (cascade should handle this, but let's be explicit)
                    # Delete all branches and their related data
                    for branch in region.branches:
                        # Delete cameras and their faults
                        for camera in branch.cameras:
                            # Delete all faults for this camera
                            Fault.query.filter_by(camera_id=camera.id).delete()
                            # Delete the camera
                            db.session.delete(camera)
                        
                        # Delete devices and their faults
                        for device in branch.devices:
                            # Delete all faults for this device
                            DeviceFault.query.filter_by(device_id=device.id).delete()
                            # Delete the device
                            db.session.delete(device)
                        
                        # Delete the branch
                        db.session.delete(branch)
                    
                    # Delete the region
                    db.session.delete(region)
                    deleted_count += 1
                    logger.info(f"Region '{region_name}' and all related data deleted by admin: {current_user.username}")
                    
            except Exception as e:
                errors.append(f"خطأ في حذف المنطقة {region_id}: {str(e)}")
                logger.error(f"Error deleting region {region_id}: {str(e)}")
                db.session.rollback()
        
        if errors:
            flash(f"تم حذف {deleted_count} مناطق بنجاح، ولكن حدثت أخطاء: {'; '.join(errors)}", "warning")
        else:
            flash(f"تم حذف {deleted_count} مناطق بنجاح مع جميع البيانات المرتبطة بها", "success")
        
        db.session.commit()
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error in delete_multiple_regions: {str(e)}", exc_info=True)
        flash(_("حدث خطأ أثناء حذف المناطق المحددة"), "danger")
    
    return redirect(url_for("regions"))

@app.route("/regions/delete/<int:region_id>", methods=["POST"])
@login_required
def delete_region(region_id):
    if current_user.role != "admin":
        flash(_("فقط المدير يمكنه حذف المناطق ❌"), "danger")
        return redirect(url_for("dashboard"))
    region = Region.query.get_or_404(region_id)
    db.session.delete(region)
    db.session.commit()
    flash("تم حذف المنطقة بنجاح ✅", "success")
    return redirect(url_for("regions"))

@app.route("/regions/import-excel", methods=["GET", "POST"])
@app.route("/chains/<int:chain_id>/regions/import-excel", methods=["GET", "POST"])
@login_required
def import_regions_excel(chain_id=None):
    """استيراد المناطق من ملف Excel"""
    if current_user.role != "admin":
        flash(_("فقط المدير يمكنه استيراد المناطق من Excel ❌"), "danger")
        return redirect(url_for("dashboard"))
    
    if request.method == "POST":
        if "excel_file" not in request.files:
            flash(t('لم يتم اختيار ملف ❌'), "danger")
            return redirect(request.url)
        
        file = request.files["excel_file"]
        
        if file.filename == "":
            flash(t('لم يتم اختيار ملف ❌'), "danger")
            return redirect(request.url)
        
        if not file.filename.endswith((".xlsx", ".xls")):
            flash(t('يجب أن يكون الملف من نوع Excel (.xlsx أو .xls) ❌'), "danger")
            return redirect(request.url)
        
        try:
            import time
            start_time = time.time()
            
            # قراءة الملف باستخدام pandas
            file_bytes = file.read()
            
            try:
                import pandas as pd
                df = pd.read_excel(io.BytesIO(file_bytes), header=0)
                logger.info(f"📊 تم قراءة {len(df)} صف من ملف Excel باستخدام pandas في {time.time() - start_time:.2f} ثانية")
            except Exception as e:
                logger.warning(f"فشل القراءة بـ pandas: {e}، استخدام openpyxl بدلاً منه")
                workbook = load_workbook(io.BytesIO(file_bytes))
                worksheet = workbook.active
                df = pd.DataFrame(worksheet.iter_rows(min_row=2, values_only=True))
            
            # التحقق من البيانات
            if df.empty:
                flash(t('الملف فارغ أو لا يحتوي على بيانات صالحة ❌'), "danger")
                return redirect(request.url)
            
            # التحقق من وجود chain_id في parameters أو URL
            if chain_id is None:
                chain_id = request.args.get("chain_id", type=int)
            
            chain = None
            if chain_id:
                chain = Chain.query.get(chain_id)
                if not chain:
                    flash(f"السلسلة رقم {chain_id} غير موجودة ❌", "danger")
                    return redirect(request.url)
                logger.info(f"🔗 سيتم ربط المناطق بالسلسلة: {chain.name}")
            
            # تهيئة المتغيرات
            regions_added = 0
            updated_regions = 0
            errors = []
            skipped_regions = 0
            
            # معالجة البيانات
            for idx, row in df.iterrows():
                try:
                    row_idx = idx + 1  # رقم الصف في Excel (1-based)
                    
                    # استخراج البيانات - لاستيراد المناطق نحتاج فقط اسم المنطقة
                    # يمكن أن يكون اسم المنطقة في أي عمود (أول عمود غير فارغ)
                    region_name = None
                    
                    # البحث عن أول عمود غير فارغ لاستخدامه كاسم المنطقة
                    for col_idx in range(len(df.columns)):
                        cell_value = str(row.iloc[col_idx]).strip() if pd.notna(row.iloc[col_idx]) else ''
                        if cell_value and cell_value != 'nan':
                            region_name = cell_value
                            break
                    
                    logger.info(f"معالجة الصف {row_idx}: منطقة='{region_name}'")
                    
                    if not region_name:
                        errors.append(f"الصف {row_idx}: اسم المنطقة إلزامي")
                        continue
                    
                    # التحقق من وجود المنطقة مسبقاً - السماح بنفس الاسم في سلاسل مختلفة
                    # البحث عن منطقة بنفس الاسم في السلسلة الحالية فقط
                    if chain_id:
                        existing_region = Region.query.filter_by(name=region_name, chain_id=chain_id).first()
                    else:
                        existing_region = Region.query.filter_by(name=region_name, chain_id=None).first()
                    
                    if existing_region:
                        # المنطقة موجودة بالفعل في نفس السياق (سلسلة محددة أو بدون سلسلة)
                        if chain:
                            logger.warning(f"⚠️ المنطقة '{region_name}' موجودة بالفعل في سلسلة '{chain.name}'")
                        else:
                            logger.warning(f"⚠️ المنطقة '{region_name}' موجودة بالفعل بدون سلسلة")
                        skipped_regions += 1
                    else:
                        # إنشاء المنطقة الجديدة مع ربطها بالسلسلة
                        region = Region(name=region_name, chain_id=chain_id if chain_id else None)
                        db.session.add(region)
                        regions_added += 1
                        
                        if chain:
                            logger.info(f"✅ تمت إضافة المنطقة: '{region_name}' لسلسلة '{chain.name}'")
                        else:
                            logger.info(f"✅ تمت إضافة المنطقة: '{region_name}' بدون سلسلة")
                
                except Exception as e:
                    errors.append(f"الصف {row_idx}: {str(e)}")
                    logger.error(f"خطأ في الصف {row_idx}: {str(e)}")
            
            # حفظ التغييرات
            try:
                commit_start = time.time()
                db.session.commit()
                commit_time = time.time() - commit_start
                
                total_time = time.time() - start_time
                
                logger.info(f"🚀 اكتملت المعالجة في {total_time:.2f} ثانية")
                logger.info(f"💾 تم حفظ التغييرات في {commit_time:.2f} ثانية")
                
                if chain:
                    if regions_added > 0 and updated_regions > 0:
                        flash(f"✅ تم إضافة {regions_added} منطقة جديدة وتحديث {updated_regions} منطقة موجودة لسلسلة '{chain.name}' من ملف Excel", "success")
                    elif regions_added > 0:
                        flash(f"✅ تم إضافة {regions_added} منطقة بنجاح لسلسلة '{chain.name}' من ملف Excel", "success")
                    elif updated_regions > 0:
                        flash(f"🔄 تم تحديث {updated_regions} منطقة لسلسلة '{chain.name}' من ملف Excel", "success")
                else:
                    if regions_added > 0 and updated_regions > 0:
                        flash(f"✅ تم إضافة {regions_added} منطقة جديدة وتحديث {updated_regions} منطقة موجودة من ملف Excel", "success")
                    elif regions_added > 0:
                        flash(f"✅ تم إضافة {regions_added} منطقة بنجاح من ملف Excel", "success")
                    elif updated_regions > 0:
                        flash(f"🔄 تم تحديث {updated_regions} منطقة من ملف Excel", "success")
                
                flash(f"⚡ تمت المعالجة بسرعة فائقة ({total_time:.2f} ثانية)", "success")
                
                if skipped_regions > 0:
                    flash(f"ℹ️ تم تجاهل {skipped_regions} منطقة موجودة بالفعل", "info")
                
                if errors:
                    logger.warning(f"أخطاء أثناء الاستيراد: {errors}")
                    flash(f"⚠️ هناك {len(errors)} أخطاء: " + " | ".join(errors[:3]), "warning")
                
                # إعادة التوجيه بناءً على السلسلة
                if chain:
                    return redirect(url_for("chain_regions", chain_id=chain.id))
                else:
                    return redirect(url_for("regions"))
            
            except Exception as e:
                db.session.rollback()
                logger.error(f"خطأ في حفظ البيانات: {str(e)}")
                flash(f"خطأ في حفظ البيانات: {str(e)} ❌", "danger")
                return redirect(request.url)
        
        except Exception as e:
            logger.error(f"خطأ في قراءة الملف: {str(e)}")
            flash(f"خطأ في قراءة الملف: {str(e)} ❌", "danger")
            return redirect(request.url)
    
    return redirect(url_for("regions"))

@app.route("/branches/<int:region_id>", methods=["GET", "POST"])
@login_required
def branches(region_id):
    if current_user.role != "admin":
        flash(_("فقط المدير يمكنه الوصول لهذه الصفحة ❌"), "danger")
        return redirect(url_for("dashboard"))
    
    # Handle POST requests (redirect to GET)
    if request.method == "POST":
        return redirect(url_for("branches", region_id=region_id))
    
    region = Region.query.get_or_404(region_id)
    # Add sequential numbering in ascending order (oldest first)
    branches_list = Branch.query.filter_by(region_id=region_id).order_by(Branch.id.asc()).all()
    branches_with_numbers = []
    for index, branch in enumerate(branches_list, 1):
        branches_with_numbers.append({
            'branch': branch,
            'row_number': index
        })
    return render_template("branches.html", branches=branches_with_numbers, region=region)

@app.route("/branches/add/<int:region_id>", methods=["GET", "POST"])
@login_required
def add_branch(region_id):
    if current_user.role != "admin":
        flash(_("فقط المدير يمكنه إضافة فروع ❌"), "danger")
        return redirect(url_for("dashboard"))
    if request.method == "POST":
        # Check if this is a North Coast region
        region = Region.query.get(region_id)
        is_north_coast = region and 'North Coast' in region.name
        
        branch = Branch(
            name=request.form['name'],
            location=request.form['location'],
            phone_number=request.form.get('phone_number', ''),
            phone_number_2=request.form.get('phone_number_2', ''),
            phone_number_3=request.form.get('phone_number_3', ''),
            phone_number_4=request.form.get('phone_number_4', ''),
            ip_address=request.form.get('ip_address', ''),
            nvr_count=int(request.form.get('nvr_count', 0)),
            branch_type='موسمي' if is_north_coast else request.form.get('branch_type', 'مؤقت'),
            region_id=region_id
        )
        db.session.add(branch)
        db.session.commit()
        print("DEBUG: Successfully added branch")
        flash("تم إضافة الفرع بنجاح ✅", "success")
        return redirect(url_for("branches", region_id=region_id))
    return render_template("add_edit_branch.html", branch=None, region=Region.query.get(region_id))

@app.route("/branches/edit/<int:branch_id>", methods=["GET", "POST"])
@login_required
def edit_branch(branch_id):
    print(f"DEBUG: edit_branch called with branch_id: {branch_id}")
    
    if current_user.role != "admin":
        print("DEBUG: User is not admin, redirecting")
        flash(_("فقط المدير يمكنه تعديل الفروع ❌"), "danger")
        return redirect(url_for("dashboard"))
    
    branch = Branch.query.get_or_404(branch_id)
    print(f"DEBUG: Found branch: {branch.name}")
    
    if request.method == "POST":
        print("DEBUG: POST request received")
        try:
            # Print all form data for debugging
            print("DEBUG: Form data received:")
            for key, value in request.form.items():
                print(f"  {key}: '{value}'")
            
            # Get form data safely
            name = request.form.get('name', '').strip()
            print(f"DEBUG: Name field: '{name}'")
            
            if not name:
                print("DEBUG: Name is empty, showing error")
                flash("اسم الفرع مطلوب ❌", "danger")
                return render_template("add_edit_branch.html", branch=branch, region=branch.region)
            
            # Update branch
            print("DEBUG: Updating branch data...")
            branch.name = name
            branch.location = request.form.get('location', '')
            branch.phone_number = request.form.get('phone_number', '')
            branch.phone_number_2 = request.form.get('phone_number_2', '')
            branch.phone_number_3 = request.form.get('phone_number_3', '')
            branch.phone_number_4 = request.form.get('phone_number_4', '')
            branch.ip_address = request.form.get('ip_address', '')
            
            # Handle nvr_count safely
            try:
                branch.nvr_count = int(request.form.get('nvr_count', 0))
            except (ValueError, TypeError):
                branch.nvr_count = 0
            
            branch.branch_type = request.form.get('branch_type', 'دائم')
            
            print("DEBUG: Attempting to commit to database...")
            db.session.commit()
            print("DEBUG: Database commit successful")
            
            flash("تم تعديل الفرع بنجاح ✅", "success")
            print("DEBUG: Redirecting to branches page")
            return redirect(url_for("branches", region_id=branch.region_id))
            
        except Exception as e:
            db.session.rollback()
            error_msg = f"ERROR in edit_branch: {str(e)}"
            print(error_msg)
            import traceback
            traceback.print_exc()
            flash(f"حدث خطأ أثناء الحفظ: {str(e)} ❌", "danger")
            return render_template("add_edit_branch.html", branch=branch, region=branch.region)
    
    print(f"DEBUG: Rendering edit form for branch: {branch.name}")
    return render_template("add_edit_branch.html", branch=branch, region=branch.region)

@app.route("/branches/delete/<int:branch_id>", methods=["POST"])
@login_required
def delete_branch(branch_id):
    print(f"DEBUG: delete_branch called with branch_id: {branch_id}")
    print(f"DEBUG: Request method: {request.method}")
    print(f"DEBUG: Request headers: {dict(request.headers)}")
    
    if current_user.role != "admin":
        print("DEBUG: User is not admin")
        flash(_("فقط المدير يمكنه حذف الفروع ❌"), "danger")
        return redirect(url_for("dashboard"))
    
    branch = Branch.query.get_or_404(branch_id)
    region_id = branch.region_id
    print(f"DEBUG: Found branch: {branch.name}")
    
    try:
        # First, delete all BBM faults associated with cameras in this branch
        cameras_in_branch = Camera.query.filter_by(branch_id=branch_id).all()
        print(f"DEBUG: Found {len(cameras_in_branch)} cameras in branch")
        
        for camera in cameras_in_branch:
            BBMFault.query.filter_by(camera_id=camera.id).delete()
        
        # Then delete the branch (this will also delete cameras due to cascade)
        print("DEBUG: Attempting to delete branch...")
        db.session.delete(branch)
        print("DEBUG: Attempting to commit to database...")
        db.session.commit()
        print("DEBUG: Database commit successful")
        
        flash("تم حذف الفرع بنجاح ✅", "success")
        print("DEBUG: Redirecting to branches page")
        
    except Exception as e:
        print(f"DEBUG: Exception occurred: {str(e)}")
        import traceback
        traceback.print_exc()
        db.session.rollback()
        print(f"DEBUG: Database rollback completed")
        flash(f"حدث خطأ أثناء حذف الفرع: {str(e)} ❌", "danger")
    
    return redirect(url_for("branches", region_id=region_id))

@app.route("/branches/delete-all/<int:region_id>", methods=["POST"])
@login_required
def delete_all_branches(region_id):
    """حذف جميع الفروع في منطقة معينة"""
    if current_user.role != "admin":
        flash(_("فقط المدير يمكنه حذف الفروع ❌"), "danger")
        return redirect(url_for("dashboard"))
    
    region = Region.query.get_or_404(region_id)
    branches_count = Branch.query.filter_by(region_id=region_id).count()
    
    if branches_count == 0:
        flash("لا توجد فروع لحذفها في هذه المنطقة ℹ️", "info")
        return redirect(url_for("branches", region_id=region_id))
    
    try:
        # حذف جميع BBM faults المرتبطة بالكاميرات في هذه الفروع
        branches = Branch.query.filter_by(region_id=region_id).all()
        for branch in branches:
            cameras_in_branch = Camera.query.filter_by(branch_id=branch.id).all()
            for camera in cameras_in_branch:
                BBMFault.query.filter_by(camera_id=camera.id).delete()
        
        # حذف جميع الفروع (سيتم حذف الكاميرات والأجهزة تلقائياً بسبب cascade)
        Branch.query.filter_by(region_id=region_id).delete()
        db.session.commit()
        
        flash(f"✅ تم حذف {branches_count} فرع من منطقة '{region.name}' بنجاح", "success")
        
    except Exception as e:
        print(f"DEBUG: Exception occurred: {str(e)}")
        import traceback
        traceback.print_exc()
        db.session.rollback()
        flash(f"حدث خطأ في حذف الفروع: {str(e)} ❌", "danger")

@app.route("/close-branch/<int:branch_id>", methods=["POST"])
@login_required
@csrf.exempt
def close_branch(branch_id):
    """إغلاق فرع مع سبب واسم المبلغ"""
    # Check if this is an AJAX request
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest' or \
             request.content_type and 'application/json' in request.content_type
    
    print(f"DEBUG: close_branch called with branch_id: {branch_id}")
    print(f"DEBUG: Request method: {request.method}")
    print(f"DEBUG: Headers: {dict(request.headers)}")
    print(f"DEBUG: Is AJAX: {is_ajax}")
    
    if current_user.role != "admin":
        print("DEBUG: User is not admin")
        if is_ajax:
            return jsonify({"success": False, "message": t("غير مصرح لك بهذا الإجراء")})
        else:
            flash(t("غير مصرح لك بهذا الإجراء"), "danger")
            return redirect(url_for("dashboard"))
    
    try:
        branch = Branch.query.get(branch_id)
        if not branch:
            print(f"DEBUG: Branch {branch_id} not found")
            if is_ajax:
                return jsonify({"success": False, "message": t("الفرع غير موجود")})
            else:
                flash(t("الفرع غير موجود"), "danger")
                return redirect(url_for("dashboard"))
        
        if branch.closed:
            print(f"DEBUG: Branch {branch_id} is already closed")
            if is_ajax:
                return jsonify({"success": False, "message": t("الفرع مغلق بالفعل")})
            else:
                flash(t("الفرع مغلق بالفعل"), "danger")
                return redirect(url_for("dashboard"))
        
        # Get form data
        closure_reason = request.form.get('closure_reason', '').strip()
        reporter_name = request.form.get('reporter_name', '').strip()
        
        print(f"DEBUG: Form data - closure_reason: '{closure_reason}', reporter_name: '{reporter_name}'")
        
        if not closure_reason or not reporter_name:
            print("DEBUG: Missing closure_reason or reporter_name")
            if is_ajax:
                return jsonify({"success": False, "message": t("يجب إدخال سبب الإغلاق واسم المبلغ")})
            else:
                flash(t("يجب إدخال سبب الإغلاق واسم المبلغ"), "danger")
                return redirect(url_for("dashboard"))
        
        # Update branch
        branch.closed = True
        branch.closure_reason = closure_reason
        branch.reporter_name = reporter_name
        branch.closure_date = utc_now()
        
        # Record in history
        history_entry = BranchHistory(
            branch_id=branch_id,
            action='close',
            action_date=utc_now(),
            reason=closure_reason,
            reporter_name=reporter_name
        )
        db.session.add(history_entry)
        
        print("DEBUG: Attempting to commit to database...")
        db.session.commit()
        print("DEBUG: Database commit successful")
        
        logger.info(f"Branch '{branch.name}' (ID: {branch_id}) closed by {reporter_name}. Reason: {closure_reason}")
        
        response_data = {
            "success": True, 
            "message": t("تم إغلاق الفرع بنجاح")
        }
        print(f"DEBUG: Returning response: {response_data}")
        response = jsonify(response_data)
        response.headers['Content-Type'] = 'application/json'
        return response
        
    except Exception as e:
        print(f"DEBUG: Exception occurred: {str(e)}")
        import traceback
        traceback.print_exc()
        logger.error(f"Error closing branch {branch_id}: {e}")
        db.session.rollback()
        
        error_response = {
            "success": False, 
            "message": t("حدث خطأ في إغلاق الفرع")
        }
        print(f"DEBUG: Returning error response: {error_response}")
        response = jsonify(error_response)
        response.headers['Content-Type'] = 'application/json'
        return response

@app.route("/reopen-branch/<int:branch_id>", methods=["POST"])
@login_required
@csrf.exempt
def reopen_branch(branch_id):
    print(f"DEBUG: reopen_branch called with branch_id: {branch_id}")
    if current_user.role != "admin":
        return jsonify({"success": False, "message": t("غير مصرح لك بهذا الإجراء")})
    
    try:
        branch = Branch.query.get(branch_id)
        if not branch:
            return jsonify({"success": False, "message": t("الفرع غير موجود")})
        
        if not branch.closed:
            return jsonify({"success": False, "message": t("الفرع مفتوح بالفعل")})
        
        # Record in history before reopening
        history_entry = BranchHistory(
            branch_id=branch_id,
            action='open',
            action_date=utc_now(),
            reason=f"Reopened from closed state. Previous reason: {branch.closure_reason or 'N/A'}",
            reporter_name=current_user.username
        )
        db.session.add(history_entry)
        
        # Update branch
        branch.closed = False
        branch.closure_reason = None
        branch.reporter_name = None
        branch.closure_date = None
        
        db.session.commit()
        
        logger.info(f"Branch '{branch.name}' (ID: {branch_id}) reopened by {current_user.username}")
        
        return jsonify({
            "success": True, 
            "message": t("تم إعادة فتح الفرع بنجاح")
        })
        
    except Exception as e:
        logger.error(f"Error reopening branch {branch_id}: {e}")
        db.session.rollback()
        return jsonify({"success": False, "message": t("حدث خطأ في إعادة فتح الفرع")})

@app.route("/closed-branches")
@login_required
def closed_branches():
    """عرض جميع الفروع المغلقة"""
    if current_user.role != "admin":
        flash(t("غير مصرح لك بالوصول إلى هذه الصفحة"), "danger")
        return redirect(url_for("dashboard"))
    
    # Get all closed branches with row numbers
    closed_branches_query = Branch.query.filter_by(closed=True).order_by(Branch.closure_date.desc())
    closed_branches = []
    
    for row_number, branch in enumerate(closed_branches_query.all(), 1):
        closed_branches.append({
            'branch': branch,
            'row_number': row_number
        })
    
    return render_template("closed_branches.html", branches=closed_branches)

@app.route("/branches/import-excel/<int:region_id>", methods=["GET", "POST"])
@login_required
def import_branches_excel(region_id):
    """استيراد الفروع من ملف Excel"""
    if current_user.role != "admin":
        flash(_("فقط المدير يمكنه استيراد الفروع من Excel ❌"), "danger")
        return redirect(url_for("dashboard"))
    
    region = Region.query.get_or_404(region_id)
    
    if request.method == "POST":
        if "excel_file" not in request.files:
            flash(t('لم يتم اختيار ملف ❌'), "danger")
            return redirect(request.url)
        
        file = request.files["excel_file"]
        
        if file.filename == "":
            flash(t('لم يتم اختيار ملف ❌'), "danger")
            return redirect(request.url)
        
        if not file.filename.endswith((".xlsx", ".xls")):
            flash(t('يجب أن يكون الملف من نوع Excel (.xlsx أو .xls) ❌'), "danger")
            return redirect(request.url)
        
        try:
            # قراءة الملف
            file_bytes = file.read()
            workbook = load_workbook(io.BytesIO(file_bytes))
            worksheet = workbook.active
            
            # التحقق من خيار التعامل مع المكررات
            handle_duplicates = request.form.get('handle_duplicates') == 'true'
            
            branches_added = 0
            cameras_added = 0
            errors = []
            processed_names = set()  # لتتبع الأسماء التي تمت معالجتها
            
            # جمع البيانات أولاً لمعالجة الكاميرات المتعددة لكل فرع
            branch_data = {}  # {branch_name: {'branch_info': {...}, 'cameras': [...]}}
            
            # القراءة من الصف الثاني (الصف الأول هو الرؤوس)
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # استخراج اسم الفرع من العمود C (العمود الثالث)
                    branch_name = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ''
                    
                    # التحقق من وجود اسم الفرع
                    if not branch_name:
                        continue
                    
                    # استخراج رقم التليفون من العمود D (العمود الرابع)
                    phone1 = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ''
                    
                    # استخراج عنوان IP من العمود B (العمود الثاني)
                    ip_address = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''
                    
                    # استخراج اسم الكاميرا من العمود A (العمود الأول)
                    camera_name = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ''
                    
                    # الموقع سيكون فارغاً لأنه غير موجود في التخطيط الجديد
                    location = ''
                    
                    # تجميع البيانات حسب اسم الفرع
                    if branch_name not in branch_data:
                        branch_data[branch_name] = {
                            'phone1': phone1,
                            'ip_address': ip_address,
                            'location': location,
                            'cameras': []
                        }
                    
                    # إضافة الكاميرا إذا كان الاسم موجود
                    if camera_name:
                        branch_data[branch_name]['cameras'].append({
                            'name': camera_name,
                            'ip_address': ip_address
                        })
                
                except Exception as e:
                    errors.append(f"الصف {row_idx}: {str(e)}")
                    logger.error(f"خطأ في الصف {row_idx}: {str(e)}")
            
            # معالجة بيانات الفروع والكاميرات
            for branch_name, data in branch_data.items():
                try:
                    # التعامل مع المكررات إذا كان الخيار مفعلاً
                    if handle_duplicates:
                        if branch_name in processed_names:
                            logger.info(f"تخطي الفرع المكرر: '{branch_name}'")
                            continue
                        processed_names.add(branch_name)
                    
                    # التحقق من وجود فرع بنفس الاسم في نفس المنطقة
                    existing_branch = Branch.query.filter_by(name=branch_name, region_id=region_id).first()
                    if existing_branch:
                        # إذا كان الفرع موجود، قم بتحديث البيانات فقط
                        if data['phone1']:
                            existing_branch.phone_number = data['phone1']
                        if data['ip_address']:
                            existing_branch.ip_address = data['ip_address']
                        
                        # إضافة/تحديث جميع الكاميرات
                        for camera_info in data['cameras']:
                            # البحث عن كاميرا بنفس الاسم في نفس الفرع
                            existing_camera = Camera.query.filter_by(name=camera_info['name'], branch_id=existing_branch.id).first()
                            if existing_camera:
                                # تحديث IP الكاميرا الموجودة
                                if camera_info['ip_address']:
                                    existing_camera.ip_address = camera_info['ip_address']
                                logger.info(f"تم تحديث الكاميرا الموجودة: '{camera_info['name']}' للفرع '{branch_name}'")
                            else:
                                # إنشاء كاميرا جديدة
                                new_camera = Camera(
                                    name=camera_info['name'],
                                    ip_address=camera_info['ip_address'] if camera_info['ip_address'] else '',
                                    camera_type='مراقبة',
                                    branch_id=existing_branch.id
                                )
                                db.session.add(new_camera)
                                cameras_added += 1
                                logger.info(f"تم إضافة كاميرا جديدة: '{camera_info['name']}' للفرع '{branch_name}'")
                        
                        logger.info(f"تم تحديث بيانات الفرع الموجود: '{branch_name}'")
                    else:
                        # الحصول على الترقيم التلقائي التالي
                        last_sequence = db.session.query(db.func.max(Branch.sequence_number)).scalar() or 0
                        next_sequence = last_sequence + 1
                        
                        # Check if this is a North Coast region
                        is_north_coast = region and 'North Coast' in region.name
                        
                        # إنشاء الفرع الجديد مع البيانات المتاحة
                        branch = Branch(
                            name=branch_name,
                            location=data['location'],
                            region_id=region_id,
                            phone_number=data['phone1'],
                            ip_address=data['ip_address'],
                            nvr_count=0,
                            sequence_number=next_sequence,
                            branch_type='موسمي' if is_north_coast else 'دائم'
                        )
                        
                        db.session.add(branch)
                        db.session.flush()  # للحصول على ID الفرع الجديد
                        
                        # إضافة جميع الكاميرات
                        for camera_info in data['cameras']:
                            new_camera = Camera(
                                name=camera_info['name'],
                                ip_address=camera_info['ip_address'] if camera_info['ip_address'] else '',
                                camera_type='مراقبة',
                                branch_id=branch.id
                            )
                            db.session.add(new_camera)
                            cameras_added += 1
                            logger.info(f"تم إضافة كاميرا جديدة: '{camera_info['name']}' للفرع الجديد '{branch_name}'")
                        
                        branches_added += 1
                
                except Exception as e:
                    errors.append(f"خطأ في معالجة الفرع '{branch_name}': {str(e)}")
                    logger.error(f"خطأ في معالجة الفرع '{branch_name}': {str(e)}")
            
            # حفظ التغييرات
            try:
                db.session.commit()
                flash(f"✅ تم إضافة {branches_added} فرع و {cameras_added} كاميرا بنجاح من ملف Excel", "success")
                
                if errors:
                    logger.warning(f"أخطاء أثناء الاستيراد: {errors}")
                    flash(f"⚠️ هناك {len(errors)} أخطاء: " + " | ".join(errors[:3]), "warning")
                
                return redirect(url_for("branches", region_id=region_id))
            
            except Exception as e:
                db.session.rollback()
                logger.error(f"خطأ في حفظ البيانات: {str(e)}")
                flash(f"خطأ في حفظ البيانات: {str(e)} ❌", "danger")
                return redirect(request.url)
        
        except Exception as e:
            logger.error(f"خطأ في قراءة الملف: {str(e)}")
            flash(f"خطأ في قراءة الملف: {str(e)} ❌", "danger")
            return redirect(request.url)
    
    return redirect(url_for("branches", region_id=region_id))

@app.route("/branches/import-phone-numbers/<int:region_id>", methods=["GET", "POST"])
@login_required
def import_phone_numbers_excel(region_id):
    """استيراد أرقام التليفونات من ملف Excel للفروع الموجودة"""
    if current_user.role != "admin":
        flash(_("فقط المدير يمكنه استيراد أرقام التليفونات ❌"), "danger")
        return redirect(url_for("dashboard"))
    
    region = Region.query.get_or_404(region_id)
    
    if request.method == "POST":
        if "excel_file" not in request.files:
            flash(t('لم يتم اختيار ملف ❌'), "danger")
            return redirect(request.url)
        
        file = request.files["excel_file"]
        
        if file.filename == "":
            flash(t('لم يتم اختيار ملف ❌'), "danger")
            return redirect(request.url)
        
        if not file.filename.endswith((".xlsx", ".xls")):
            flash(t('يجب أن يكون الملف من نوع Excel (.xlsx أو .xls) ❌'), "danger")
            return redirect(request.url)
        
        try:
            # قراءة الملف
            file_bytes = file.read()
            workbook = load_workbook(io.BytesIO(file_bytes))
            worksheet = workbook.active
            
            branches_updated = 0
            errors = []
            not_found_branches = []
            
            # القراءة من الصف الثاني (الصف الأول هو الرؤوس)
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # استخراج اسم الفرع من العمود C (العمود الثالث)
                    branch_name = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ''
                    
                    # استخراج رقم التليفون من العمود D (العمود الرابع)
                    phone_number = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ''
                    
                    # التحقق من وجود اسم الفرع ورقم التليفون
                    if not branch_name or not phone_number:
                        continue
                    
                    # البحث عن الفرع في نفس المنطقة
                    existing_branch = Branch.query.filter_by(name=branch_name, region_id=region_id).first()
                    if existing_branch:
                        # تحديث رقم التليفون الأول للفرع
                        existing_branch.phone_number = phone_number
                        branches_updated += 1
                        logger.info(f"تم تحديث رقم التليفون للفرع: '{branch_name}' بالرقم: {phone_number}")
                    else:
                        not_found_branches.append(branch_name)
                        logger.warning(f"الفرع '{branch_name}' غير موجود في المنطقة '{region.name}'")
                
                except Exception as e:
                    errors.append(f"الصف {row_idx}: {str(e)}")
                    logger.error(f"خطأ في الصف {row_idx}: {str(e)}")
            
            # حفظ التغييرات
            try:
                db.session.commit()
                flash(f"✅ تم تحديث أرقام التليفونات لـ {branches_updated} فرع بنجاح", "success")
                
                if not_found_branches:
                    flash(f"⚠️ فروع غير موجودة: {', '.join(not_found_branches[:5])}" + 
                          (f" و {len(not_found_branches) - 5} أخرى" if len(not_found_branches) > 5 else ""), "warning")
                
                if errors:
                    logger.warning(f"أخطاء أثناء الاستيراد: {errors}")
                    flash(f"⚠️ هناك {len(errors)} أخطاء: " + " | ".join(errors[:3]), "warning")
                
                return redirect(url_for("branches", region_id=region_id))
            
            except Exception as e:
                db.session.rollback()
                logger.error(f"خطأ في حفظ البيانات: {str(e)}")
                flash(f"خطأ في حفظ البيانات: {str(e)} ❌", "danger")
                return redirect(request.url)
        
        except Exception as e:
            logger.error(f"خطأ في قراءة الملف: {str(e)}")
            flash(f"خطأ في قراءة الملف: {str(e)} ❌", "danger")
            return redirect(request.url)
    
    return redirect(url_for("branches", region_id=region_id))

@app.route("/cameras/<int:branch_id>")
@login_required
def cameras(branch_id):
    if current_user.role != "admin":
        flash(_("فقط المدير يمكنه الوصول لهذه الصفحة ❌"), "danger")
        return redirect(url_for("dashboard"))
    branch = Branch.query.get_or_404(branch_id)
    return render_template("cameras.html", cameras=branch.cameras, branch=branch)

@app.route("/cameras/add/<int:branch_id>", methods=["GET", "POST"])
@login_required
def add_camera(branch_id):
    if current_user.role != "admin":
        flash("فقط المدير يمكنه إضافة كاميرات ❌", "danger")
        return redirect(url_for("dashboard"))
    if request.method == "POST":
        camera = Camera(
            name=request.form['name'],
            ip_address=request.form['ip_address'],
            camera_type=request.form.get('camera_type', 'مراقبة'),
            branch_id=branch_id
        )
        db.session.add(camera)
        db.session.commit()
        flash("تم إضافة الكاميرا بنجاح ✅", "success")
        return redirect(url_for("cameras", branch_id=branch_id))
    return render_template("add_edit_camera.html", camera=None, branch=Branch.query.get(branch_id))

@app.route("/cameras/edit/<int:camera_id>", methods=["GET", "POST"])
@login_required
def edit_camera(camera_id):
    if current_user.role != "admin":
        flash("فقط المدير يمكنه تعديل الكاميرات ❌", "danger")
        return redirect(url_for("dashboard"))
    camera = Camera.query.get_or_404(camera_id)
    if request.method == "POST":
        camera.name = request.form['name']
        camera.ip_address = request.form['ip_address']
        camera.camera_type = request.form.get('camera_type', 'مراقبة')
        db.session.commit()
        flash("تم تعديل الكاميرا بنجاح ✅", "success")
        return redirect(url_for("cameras", branch_id=camera.branch_id))
    return render_template("add_edit_camera.html", camera=camera, branch=camera.branch)

@app.route("/cameras/delete/<int:camera_id>", methods=["POST"])
@login_required
def delete_camera(camera_id):
    if current_user.role != "admin":
        flash("فقط المدير يمكنه حذف الكاميرات ❌", "danger")
        return redirect(url_for("dashboard"))
    camera = Camera.query.get_or_404(camera_id)
    branch_id = camera.branch_id
    
    try:
        # First, delete all BBM faults associated with this camera
        BBMFault.query.filter_by(camera_id=camera_id).delete()
        
        # Then delete the camera
        db.session.delete(camera)
        db.session.commit()
        flash("تم حذف الكاميرا بنجاح ✅", "success")
    except Exception as e:
        db.session.rollback()
        print(f"Error deleting camera: {str(e)}")
        flash(f"حدث خطأ أثناء حذف الكاميرا: {str(e)} ❌", "danger")
    
    return redirect(url_for("cameras", branch_id=branch_id))

@app.route("/cameras/import-excel/<int:branch_id>", methods=["GET", "POST"])
@login_required
def import_cameras_excel(branch_id):
    """استيراد الكاميرات من ملف Excel مع دعم ملفات متعددة الفروع"""
    if current_user.role != "admin":
        flash("فقط المدير يمكنه استيراد الكاميرات من Excel ❌", "danger")
        return redirect(url_for("dashboard"))
    
    branch = Branch.query.get_or_404(branch_id)
    
    if request.method == "POST":
        if "excel_file" not in request.files:
            flash(t('لم يتم اختيار ملف ❌'), "danger")
            return redirect(request.url)
        
        file = request.files["excel_file"]
        
        if file.filename == "":
            flash(t('لم يتم اختيار ملف ❌'), "danger")
            return redirect(request.url)
        
        if not file.filename.endswith((".xlsx", ".xls")):
            flash(t('يجب أن يكون الملف من نوع Excel (.xlsx أو .xls) ❌'), "danger")
            return redirect(request.url)
        
        try:
            import time
            start_time = time.time()
            
            # قراءة الملف باستخدام pandas للسرعة الفائقة
            file_bytes = file.read()
            
            # محاولة القراءة باستخدام pandas أولاً
            try:
                import pandas as pd
                df = pd.read_excel(io.BytesIO(file_bytes), header=0)
                logger.info(f"📊 تم قراءة {len(df)} صف من ملف Excel باستخدام pandas في {time.time() - start_time:.2f} ثانية")
            except Exception as e:
                # الرجوع إلى openpyxl في حالة الفشل
                logger.warning(f"فشل القراءة بـ pandas: {e}، استخدام openpyxl بدلاً منه")
                workbook = load_workbook(io.BytesIO(file_bytes))
                worksheet = workbook.active
                df = pd.DataFrame(worksheet.iter_rows(min_row=2, values_only=True))
            
            # التحقق من البيانات
            if df.empty:
                flash(t('الملف فارغ أو لا يحتوي على بيانات صالحة ❌'), "danger")
                return redirect(request.url)
            
            # تهيئة المتغيرات
            cameras_added = 0
            errors = []
            skipped_cameras = 0
            notes_found = []
            ip_matches = 0
            
            # تحضير البيانات مسبقاً للسرعة
            current_branch_clean = branch.name.strip().lower()
            current_branch_clean = ' '.join(current_branch_clean.split())
            current_numbers = [word for word in current_branch_clean.split() if word.isdigit()]
            
            # جلب كاميرات الفرع الحالي مرة واحدة للتخزين المؤقت
            existing_cameras = Camera.query.filter_by(branch_id=branch_id).all()
            existing_camera_names = {cam.name.lower() for cam in existing_cameras}
            existing_ip_prefixes = set()
            
            for cam in existing_cameras:
                if cam.ip_address:
                    ip_parts = cam.ip_address.split('.')
                    if len(ip_parts) >= 3:
                        prefix = '.'.join(ip_parts[:3])
                        existing_ip_prefixes.add(prefix)
            
            # تجهيز قائمة الكاميرات للإضافة السريعة
            cameras_to_add = []
            processing_start = time.time()
            
            # معالجة البيانات بالتوازي للسرعة القصوى
            # عرض أول صف للتشخيص
            if len(df) > 0:
                first_row = df.iloc[0]
                logger.info(f"🔍 DEBUG: أول صف من البيانات (كما تم قراءته):")
                for i, col in enumerate(first_row):
                    logger.info(f"   العمود {i+1}: '{col}' (نوع: {type(col)})")
                logger.info(f"🔍 DEBUG: يبحث عن الفرع: '{current_branch_clean}'")
                logger.info(f"🔍 DEBUG: أرقام الفرع الحالي: {current_numbers}")
                logger.info(f"🔍 DEBUG: عدد الأعمدة: {len(df.columns)}")
                
                # عرض أول 5 صفوف للتأكد
                logger.info(f"🔍 DEBUG: أول 5 صفوف من البيانات:")
                for idx in range(min(5, len(df))):
                    row = df.iloc[idx]
                    logger.info(f"   الصف {idx+1}: {[str(cell) for cell in row]}")
            
            for idx, row in df.iterrows():
                try:
                    row_idx = idx + 2  # تعديل الرقم ليطابق Excel
                    
                    # استخراج البيانات مع البحث المرن عن اسم الفرع
                    if len(df.columns) >= 5:
                        col1 = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ''
                        col2 = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ''
                        col3 = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else ''
                        col4 = str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) else 'مراقبة'
                        col5 = str(row.iloc[4]).strip() if pd.notna(row.iloc[4]) else ''
                        
                        # البحث الذكي عن اسم الفرع في أي عمود
                        all_columns = [col1, col2, col3, col4, col5]
                        branch_name = ''
                        branch_found = False
                        branch_column_index = -1
                        
                        for i, col in enumerate([col1, col2, col3, col4, col5]):
                            col_clean = col.strip().lower()
                            logger.info(f"🔍 DEBUG: فحص العمود {i+1}: '{col}' | يبحث عن: {current_numbers}")
                            
                            # تطابق مرن - ابحث عن أي رقم من الفرع الحالي OR اسم الفرع
                            branch_match = False
                            
                            # البحث بالأرقام أولاً
                            if current_numbers:
                                for num in current_numbers:
                                    if num in col_clean:
                                        branch_match = True
                                        logger.info(f"✅ DEBUG: تم العثور على تطابق الرقم {num} في العمود {i+1}: '{col}'")
                                        break
                            
                            # إذا لم يتم العثور على رقم، ابحث بالاسم
                            if not branch_match:
                                # استخراج كلمات اسم الفرع الحالي (بدون أرقام)
                                branch_words = [word for word in current_branch_clean.split() if not word.isdigit()]
                                for word in branch_words:
                                    if len(word) > 2 and word in col_clean:  # تجاهل الكلمات القصيرة
                                        branch_match = True
                                        logger.info(f"✅ DEBUG: تم العثور على تطابق الاسم '{word}' في العمود {i+1}: '{col}'")
                                        break
                            
                            if branch_match:
                                branch_found = True
                                branch_column_index = i
                                break
                        
                        if branch_found:
                            logger.info(f"🎯 DEBUG: تم العثور على تطابق لفرع '{branch.name}' في العمود {branch_column_index+1}")
                            
                            # استخراج البيانات بناءً على الترتيب الثابت للأعمدة
                            camera_name = col1.strip()  # العمود A
                            ip_address = col2.strip()   # العمود B
                            branch_name = col3.strip()  # العمود C

                            logger.info(f"🎯 DEBUG: تم استخراج - كاميرا: '{camera_name}' من العمود A, IP: '{ip_address}' من العمود B, فرع: '{branch_name}' من العمود C")
                            
                            camera_type = col4.strip() if col4.strip() else 'مراقبة'
                            notes = col5.strip() if col5.strip() else ''
                            
                            if not camera_name or not ip_address:
                                logger.warning(f"⚠️ DEBUG: لم يتم العثور على اسم كاميرا أو IP صالح")
                                continue
                            
                            # قبول مباشر بعد التحقق من الفرع
                            is_match = True
                            logger.info(f"✅ تم قبول الكاميرا '{camera_name}' (ينتمي للفرع الحالي '{branch.name}')")
                        else:
                            logger.warning(f" DEBUG: لم يتم العثور على تطابق للفرع '{branch.name}' في أي عمود")
                            logger.info(f" DEBUG: الأعمدة: 1='{col1}', 2='{col2}', 3='{col3}', 4='{col4}', 5='{col5}'")
                            # تجاهل الصف إذا لم يتم العثور على تطابق للفرع
                            is_match = False
                    else:
                        # التعامل مع الأعمدة الأقل
                        col1 = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ''
                        col2 = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ''
                        col3 = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else ''
                        
                        # تهيئة المتغيرات
                        camera_name = ''
                        ip_address = ''
                        branch_name = ''
                        is_match = False
                        
                        # البحث عن الفرع في الأعمدة المتاحة
                        if current_branch_clean in col1.lower() or col1.lower() in current_branch_clean:
                            branch_name = col1
                            camera_name = col2
                            ip_address = col3
                            is_match = True
                        elif current_branch_clean in col2.lower() or col2.lower() in current_branch_clean:
                            branch_name = col2
                            camera_name = col1
                            ip_address = col3
                            is_match = True
                        else:
                            branch_name = col1
                            camera_name = col2
                            ip_address = col3
                            # تطابق مرن بالأرقام
                            if current_numbers:
                                for num in current_numbers:
                                    if num in col1.lower() or num in col2.lower() or num in col3.lower():
                                        is_match = True
                                        logger.info(f" DEBUG: تم العثور على تطابق الرقم {num} في الأعمدة القليلة")
                                        break
                        
                        # فحص ذكي وتبديل البيانات إذا كانت في الأعمدة الخاطئة
                        if is_match:
                            # التحقق إذا كان اسم الكاميرا يحتوي أرقام IP والعكس
                            if camera_name and ip_address:
                                # فحص إذا كان اسم الكاميرا يشبه عنوان IP
                                if _looks_like_ip(camera_name) and not _looks_like_ip(ip_address):
                                    logger.info(f"🔄 تبديل: اسم الكاميرا '{camera_name}' يشبه IP، IP '{ip_address}' يشبه اسم - يتم التبديل")
                                    camera_name, ip_address = ip_address, camera_name
                                elif _looks_like_ip(ip_address) and not _looks_like_ip(camera_name):
                                    logger.info(f"🔄 تبديل: IP '{ip_address}' يشبه اسم، اسم الكاميرا '{camera_name}' يشبه IP - يتم التبديل")
                                    camera_name, ip_address = camera_name, ip_address
                            
                            logger.info(f"🎯 DEBUG: تم استخراج - كاميرا: '{camera_name}' من العمود A, IP: '{ip_address}' من العمود B, فرع: '{branch_name}' من العمود C")
                        
                        camera_type = 'مراقبة'
                        notes = ''
                    
                    # التحقق من البيانات الأساسية
                    if not branch_name or not camera_name or not ip_address:
                        if not branch_name:
                            errors.append(f"الصف {row_idx}: اسم الفرع إلزامي")
                        elif not camera_name:
                            errors.append(f"الصف {row_idx}: اسم الكاميرا إلزامي")
                        else:
                            errors.append(f"الصف {row_idx}: عنوان IP إلزامي")
                        continue
                    
                    # تبديل الأعمدة تلقائياً إذا كانت البيانات في أماكن خاطئة
                    if camera_name and ip_address:
                        camera_has_digits = any(char.isdigit() for char in camera_name)
                        camera_has_alpha = any(char.isalpha() for char in camera_name)
                        ip_has_digits = any(char.isdigit() for char in ip_address)
                        ip_has_alpha = any(char.isalpha() for char in ip_address)
                        
                        if camera_has_digits and not camera_has_alpha and ip_has_alpha and not ip_has_digits:
                            camera_name, ip_address = ip_address, camera_name
                            logger.info(f"تم تبديل الأعمدة للصف {row_idx}: {camera_name} ←→ {ip_address}")
                    
                    # التحقق من وجود كاميرا بنفس الاسم
                    if camera_name.lower() in existing_camera_names:
                        errors.append(f"الصف {row_idx}: كاميرا باسم '{camera_name}' موجودة بالفعل")
                        continue
                    
                    # التحقق النهائي من تطابق الفرع
                    if not is_match:
                        logger.warning(f"❌ تم تجاهل الكاميرا '{camera_name}' - لا تنتمي للفرع '{branch.name}'")
                        skipped_cameras += 1
                        continue
                    
                    # إنشاء الكاميرا الجديدة
                    camera = Camera(
                        name=camera_name,
                        ip_address=ip_address,
                        camera_type=camera_type,
                        branch_id=branch_id
                    )
                    
                    # إضافة الملاحظات إذا وجدت
                    if notes:
                        camera.note = notes
                        logger.info(f"📝 ملاحظات للكاميرا '{camera_name}': {notes}")
                        notes_found.append(camera_name)
                    
                    db.session.add(camera)
                    cameras_added += 1
                
                except Exception as e:
                    errors.append(f"الصف {row_idx}: {str(e)}")
                    logger.error(f"خطأ في الصف {row_idx}: {str(e)}")
            
            # حفظ التغييرات مع تحسين الأداء
            try:
                commit_start = time.time()
                db.session.commit()
                commit_time = time.time() - commit_start
                
                total_time = time.time() - start_time
                processing_time = time.time() - processing_start
                
                logger.info(f"🚀 اكتملت المعالجة في {total_time:.2f} ثانية")
                logger.info(f"📊 تمت معالجة {len(df)} صف في {processing_time:.2f} ثانية")
                logger.info(f"💾 تم حفظ التغييرات في {commit_time:.2f} ثانية")
                
                flash(f"✅ تم إضافة {cameras_added} كاميرا بنجاح من ملف Excel", "success")
                flash(f"⚡ تمت المعالجة بسرعة فائقة ({processing_time:.2f} ثانية)", "success")
                
                if skipped_cameras > 0:
                    flash(f"ℹ️ تم تجاهل {skipped_cameras} كاميرا لا تخص فرع '{branch.name}'", "info")
                
                if ip_matches > 0:
                    flash(f"🔍 تم استيراد {ip_matches} كاميرا بناءً على تطابق عنوان IP فقط", "info")
                
                if notes_found:
                    flash(f"📝 تم العثور على ملاحظات في {len(notes_found)} كاميرات:", "info")
                    for note in notes_found[:3]:  # عرض أول 3 ملاحظات فقط
                        flash(f"   • {note}", "info")
                    if len(notes_found) > 3:
                        flash(f"   ... و {len(notes_found) - 3} ملاحظات أخرى", "info")
                
                if errors:
                    logger.warning(f"أخطاء أثناء الاستيراد: {errors}")
                    flash(f"⚠️ هناك {len(errors)} أخطاء: " + " | ".join(errors[:3]), "warning")
                
                return redirect(url_for("cameras", branch_id=branch_id))
            
            except Exception as e:
                db.session.rollback()
                logger.error(f"خطأ في حفظ البيانات: {str(e)}")
                flash(f"خطأ في حفظ البيانات: {str(e)} ❌", "danger")
                return redirect(request.url)
        
        except Exception as e:
            logger.error(f"خطأ في قراءة الملف: {str(e)}")
            flash(f"خطأ في قراءة الملف: {str(e)} ❌", "danger")
            return redirect(request.url)
    
    return redirect(url_for("cameras", branch_id=branch_id))

@app.route("/cameras/delete-all/<int:branch_id>", methods=["POST"])
@login_required
def delete_all_cameras(branch_id):
    """حذف جميع الكاميرات في فرع معين"""
    if current_user.role != "admin":
        flash(_("فقط المدير يمكنه حذف الكاميرات ❌"), "danger")
        return redirect(url_for("dashboard"))
    
    branch = Branch.query.get_or_404(branch_id)
    cameras_count = Camera.query.filter_by(branch_id=branch_id).count()
    
    if cameras_count == 0:
        flash("لا توجد كاميرات لحذفها في هذا الفرع ℹ️", "info")
        return redirect(url_for("cameras", branch_id=branch_id))
    
    try:
        # حذف جميع BBM faults المرتبطة بالكاميرات في هذا الفرع
        cameras = Camera.query.filter_by(branch_id=branch_id).all()
        for camera in cameras:
            BBMFault.query.filter_by(camera_id=camera.id).delete()
        
        # حذف جميع الكاميرات
        Camera.query.filter_by(branch_id=branch_id).delete()
        db.session.commit()
        
        flash(f"✅ تم حذف {cameras_count} كاميرا من فرع '{branch.name}' بنجاح", "success")
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"خطأ في حذف الكاميرات: {str(e)}")
        flash(f"حدث خطأ أثناء حذف الكاميرات: {str(e)} ❌", "danger")
    
    return redirect(url_for("cameras", branch_id=branch_id))

# Device Management Routes
@app.route("/devices")
@login_required
def devices():
    if current_user.role != "admin":
        flash(_("فقط المدير يمكنه الوصول لهذه الصفحة ❌"), "danger")
        return redirect(url_for("dashboard"))
    
    branch_id = request.args.get("branch_id", type=int)
    if branch_id:
        devices_list = Device.query.filter_by(branch_id=branch_id).all()
        branch = Branch.query.get(branch_id)
    else:
        devices_list = Device.query.all()
        branch = None
    
    return render_template("devices.html", devices=devices_list, branch=branch)

@app.route("/devices/add/<int:branch_id>", methods=["GET", "POST"])
@login_required
def add_device(branch_id):
    if current_user.role != "admin":
        flash("فقط المدير يمكنه إضافة أجهزة ❌", "danger")
        return redirect(url_for("dashboard"))
    
    branch = Branch.query.get_or_404(branch_id)
    
    if request.method == "POST":
        device = Device(
            name=request.form['name'],
            device_type=request.form['device_type'],
            ip_address=request.form.get('ip_address', ''),
            location=request.form.get('location', ''),
            branch_id=branch_id
        )
        db.session.add(device)
        db.session.commit()
        flash("تم إضافة الجهاز بنجاح ✅", "success")
        return redirect(url_for("devices", branch_id=branch_id))
    
    return render_template("add_edit_device.html", device=None, branch=branch)

@app.route("/devices/edit/<int:device_id>", methods=["GET", "POST"])
@login_required
def edit_device(device_id):
    if current_user.role != "admin":
        flash("فقط المدير يمكنه تعديل الأجهزة ❌", "danger")
        return redirect(url_for("dashboard"))
    
    device = Device.query.get_or_404(device_id)
    
    if request.method == "POST":
        device.name = request.form['name']
        device.device_type = request.form['device_type']
        device.ip_address = request.form.get('ip_address', '')
        device.location = request.form.get('location', '')
        db.session.commit()
        flash("تم تعديل الجهاز بنجاح ✅", "success")
        return redirect(url_for("devices", branch_id=device.branch_id))
    
    return render_template("add_edit_device.html", device=device, branch=device.branch)

@app.route("/devices/delete/<int:device_id>", methods=["POST"])
@login_required
def delete_device(device_id):
    if current_user.role != "admin":
        flash("فقط المدير يمكنه حذف الأجهزة ❌", "danger")
        return redirect(url_for("dashboard"))
    
    device = Device.query.get_or_404(device_id)
    branch_id = device.branch_id
    db.session.delete(device)
    db.session.commit()
    flash("تم حذف الجهاز بنجاح ✅", "success")
    return redirect(url_for("devices", branch_id=branch_id))

# Device Faults Routes
@app.route("/device-faults")
@login_required
def all_device_faults():
    if current_user.role == "admin":
        device_faults = DeviceFault.query.order_by(DeviceFault.date_reported.desc()).all()
    else:
        device_faults = DeviceFault.query.filter(
            DeviceFault.technician_id == current_user.id
        ).order_by(DeviceFault.date_reported.desc()).all()
    return render_template("all_device_faults.html", device_faults=device_faults)

@app.route("/device-faults/<int:device_id>")
@login_required
def device_faults(device_id):
    device = Device.query.get_or_404(device_id)
    technicians = User.query.filter_by(role='technician').all()
    now = utc_now()

    if current_user.role != "admin":
        flash(_("فقط المدير يمكنه الوصول لهذه الصفحة ❌"), "danger")
        return redirect(url_for("all_device_faults"))

    # compute average repair duration (in days) for resolved faults of this device
    faults_list = list(device.faults)
    resolved = [f for f in faults_list if f.resolved_at]
    if resolved:
        total_seconds = sum((f.resolved_at - f.date_reported).total_seconds() for f in resolved)
        avg_days = total_seconds / len(resolved) / 86400.0
    else:
        avg_days = 0

    return render_template("device_faults.html", faults=faults_list, device=device, technicians=technicians, now=now, avg_days=avg_days)

@app.route("/device-faults/add/<int:device_id>", methods=["GET", "POST"])
@login_required
def add_device_fault(device_id):
    if current_user.role != "admin":
        flash("فقط المدير يمكنه إضافة أعطال ❌", "danger")
        return redirect(url_for("all_device_faults"))

    device = Device.query.get_or_404(device_id)
    technicians = User.query.filter_by(role='technician').all()

    if request.method == "POST":
        technician_id = request.form.get('technician_id')
        technician_id = int(technician_id) if technician_id else None
        fault = DeviceFault(
            description=request.form['description'],
            fault_type=request.form['fault_type'],
            reported_by=request.form['reported_by'],
            technician_id=technician_id,
            device_id=device_id
        )
        db.session.add(fault)
        db.session.commit()
        flash("تم تسجيل عطل الجهاز بنجاح ✅", "success")
        return redirect(url_for("device_faults", device_id=device_id))

    return render_template("add_edit_device_fault.html", fault=None, device=device, technicians=technicians)

@app.route("/device-faults/edit/<int:fault_id>", methods=["GET", "POST"])
@login_required
def edit_device_fault(fault_id):
    if current_user.role != "admin":
        flash("فقط المدير يمكنه تعديل الأعطال ❌", "danger")
        return redirect(url_for("all_device_faults"))

    fault = DeviceFault.query.get_or_404(fault_id)
    technicians = User.query.filter_by(role='technician').all()

    if request.method == "POST":
        technician_id = request.form.get('technician_id')
        technician_id = int(technician_id) if technician_id else None
        
        fault.description = request.form['description']
        fault.fault_type = request.form['fault_type']
        fault.reported_by = request.form['reported_by']
        fault.technician_id = technician_id
        
        db.session.commit()
        flash("تم تعديل عطل الجهاز بنجاح ✅", "success")
        return redirect(url_for("device_faults", device_id=fault.device_id))

    return render_template("add_edit_device_fault.html", fault=fault, device=fault.device, technicians=technicians)

@app.route("/device-faults/resolve/<int:fault_id>", methods=["GET", "POST"])
@login_required
def resolve_device_fault(fault_id):
    fault = DeviceFault.query.get_or_404(fault_id)

    if current_user.role == "technician" and fault.technician_id != current_user.id:
        flash("غير مسموح لك بإصلاح هذا العطل ❌", "danger")
        return redirect(url_for("all_device_faults"))

    if request.method == "POST":
        fault.resolved_at = utc_now()
        fault.resolved_by = current_user.username
        fault.repair_notes = request.form.get('repair_notes', '')

        if 'repair_image' in request.files:
            file = request.files['repair_image']
            if file and file.filename != "" and allowed_upload_file(file.filename, {'png', 'jpg', 'jpeg', 'gif'}):
                filename = secure_filename(file.filename)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(filepath)
                fault.repair_image = filename
            elif file and file.filename and not allowed_upload_file(file.filename, {'png', 'jpg', 'jpeg', 'gif'}):
                flash("صيغة الملف غير مدعومة. استخدم: " + ", ".join(app.config['ALLOWED_IMAGE_EXTENSIONS']), "warning")

        db.session.commit()
        flash("تم إصلاح عطل الجهاز بنجاح ✅", "success")
        return redirect(url_for("all_device_faults"))

    now = utc_now()
    return render_template("resolve_device_fault.html", fault=fault, now=now)

@app.route("/device-faults/details/<int:fault_id>")
@login_required
def device_fault_details(fault_id):
    fault = DeviceFault.query.get_or_404(fault_id)
    
    if current_user.role == "technician" and fault.technician_id != current_user.id:
        flash("غير مسموح لك بمشاهدة هذا العطل ❌", "danger")
        return redirect(url_for("all_device_faults"))

    now = utc_now()
    return render_template("device_fault_details.html", fault=fault, now=now)

@app.route("/device-faults/delete/<int:fault_id>", methods=["POST"])
@login_required
@csrf.exempt
def delete_device_fault(fault_id):
    if current_user.role != "admin":
        return jsonify({"success": False, "message": t("غير مصرح لك بهذا الإجراء")})
    
    fault = DeviceFault.query.get_or_404(fault_id)
    
    try:
        db.session.delete(fault)
        db.session.commit()
        
        logger.info(f"Device fault {fault_id} deleted by admin {current_user.username}")
        
        return jsonify({
            "success": True, 
            "message": t("تم حذف العطل بنجاح")
        })
        
    except Exception as e:
        logger.error(f"Error deleting device fault {fault_id}: {e}")
        db.session.rollback()
        return jsonify({"success": False, "message": t("حدث خطأ في حذف العطل")})

@app.route("/device-faults/delete-all", methods=["POST"])
@login_required
@csrf.exempt
def delete_all_device_faults():
    if current_user.role != "admin":
        return jsonify({"success": False, "message": t("غير مصرح لك بهذا الإجراء")})
    
    try:
        # Count before deletion
        count = DeviceFault.query.count()
        
        if count == 0:
            return jsonify({"success": False, "message": t("لا توجد أعطال لحذفها")})
        
        # Delete all device faults
        DeviceFault.query.delete()
        db.session.commit()
        
        logger.info(f"All {count} device faults deleted by admin {current_user.username}")
        
        return jsonify({
            "success": True, 
            "message": t("تم حذف جميع الأعطال بنجاح"),
            "deleted_count": count
        })
        
    except Exception as e:
        logger.error(f"Error deleting all device faults: {e}")
        db.session.rollback()
        return jsonify({"success": False, "message": t("حدث خطأ في حذف الأعطال")})

@app.route("/faults/details/<int:fault_id>")
@login_required
def fault_details(fault_id):
    fault = Fault.query.get_or_404(fault_id)
    

    return render_template("fault_details.html", fault=fault)

def get_menu_items(user_role="admin"):
    """Get menu items for sidebar with translations"""
    items = load_menu_items()
    lang = getattr(g, 'lang', 'en')
    
    # Add translations to menu items
    translated_items = []
    for item in items:
        translated_item = item.copy()
        # Use display_name as translation key
        translated_item['display_name'] = t(item['display_name'])
        translated_items.append(translated_item)
    
    return translated_items

# Make the function available to templates
app.jinja_env.globals['get_menu_items'] = get_menu_items

@app.route("/menu-items")
@login_required
def get_menu_items_api():
    """API endpoint to get current menu items"""
    try:
        menu_items = get_menu_items(current_user.role)
        
        # Process menu items to include translated display names
        processed_items = []
        for item in menu_items:
            processed_item = item.copy()
            # Apply translation to display name
            processed_item['display_name'] = _(item['display_name'])
            processed_items.append(processed_item)
        
        return jsonify(processed_items)
    except Exception as e:
        print(f"Error in menu-items API: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/search-camera-faults")
@login_required
def search_camera_faults():
    """Search page for camera faults."""
    if current_user.role == "admin":
        faults = Fault.query.order_by(Fault.date_reported.desc()).all()
    else:
        faults = Fault.query.filter_by(technician_id=current_user.id).order_by(Fault.date_reported.desc()).all()
    
    return render_template("search_camera_faults.html", faults=faults)
@app.route("/faults")
@login_required
def faults_redirect():
    """Redirect /faults to /all-faults for menu compatibility"""
    return redirect(url_for("all_faults"))

@app.route("/technicians")
@login_required
def technicians():
    if current_user.role != "admin":
        flash(_("فقط المدير يمكنه الوصول لهذه الصفحة ❌"), "danger")
        return redirect(url_for("dashboard"))
    technicians_list = User.query.filter(User.role == "technician").all()
    return render_template("technicians.html", technicians=technicians_list)

@app.route("/technicians/add", methods=["GET", "POST"])
@login_required
def add_technician():
    if current_user.role != "admin":
        flash("فقط المدير يمكنه إضافة فنيين ❌", "danger")
        return redirect(url_for("dashboard"))
    if request.method == "POST":
        username = request.form['username']
        password = generate_password_hash(request.form['password'])
        user = User(username=username, password=password, role="technician")
        db.session.add(user)
        db.session.commit()
        flash("تم إضافة الفني بنجاح ✅", "success")
        return redirect(url_for("technicians"))
    return render_template("add_technician.html")

@app.route("/technicians/delete/<int:tech_id>", methods=["POST"])
@login_required
def delete_technician(tech_id):
    if current_user.role != "admin":
        flash("فقط المدير يمكنه حذف فنيين ❌", "danger")
        return redirect(url_for("dashboard"))

    tech = User.query.get_or_404(tech_id)
    technician_faults = Fault.query.filter_by(technician_id=tech_id).count()
    if technician_faults > 0:
        flash(f"لا يمكن حذف الفني لأنه لديه {technician_faults} عطل مسند إليه ❌", "danger")
        return redirect(url_for("technicians"))

    db.session.delete(tech)
    db.session.commit()
    flash("تم حذف الفني بنجاح ✅", "success")
    return redirect(url_for("technicians"))

@app.route("/all-faults")
@login_required
def all_faults():
    """Main faults page - redirect based on user role"""
    if current_user.role == "admin":
        return redirect(url_for("camera_faults_sidebar"))
    else:
        return redirect(url_for("search_camera_faults"))

@app.route("/camera-faults-sidebar")
@login_required
def camera_faults_sidebar():
    """Camera faults sidebar page with interactive management"""
    if current_user.role != "admin":
        flash("فقط المدير يمكنه الوصول لهذه الصفحة ❌", "danger")
        return redirect(url_for("dashboard"))
    
    # Add translations for JavaScript
    import json
    js_translations = {
        'convertedToBPM': _('محول الي BPM'),
        'resolved': _('مُصلح'),
        'pending': _('قيد الانتظار'),
        'edit': _('تعديل'),
        'convertToBPM': _('تحويل إلى BPM'),
        'repair': _('إصلاح'),
        'view': _('عرض'),
        'delete': _('حذف'),
        'errorLoadingFaults': _('حدث خطأ في تحميل الأعطال'),
        'faultUpdatedSuccess': _('تم تعديل العطل بنجاح'),
        'errorOccurred': _('حدث خطأ'),
        'errorSavingChanges': _('حدث خطأ في حفظ التعديلات'),
        'faultRepairedSuccess': _('تم إصلاح العطل بنجاح'),
        'errorRepairingFault': _('حدث خطأ في إصلاح العطل'),
        'confirmConvertBPM': _('هل أنت متأكد من تحويل هذا العطل إلى BPM؟ سيتم نقله إلى قسم BPM ولن يظهر في قائمة الأعطال الرئيسية'),
        'faultConvertedBPM': _('تم تحويل العطل إلى BPM بنجاح'),
        'errorConvertingBPM': _('حدث خطأ في تحويل العطل إلى BPM'),
        'confirmDelete': _('هل أنت متأكد من حذف هذا العطل؟ لا يمكن التراجع عن هذا الإجراء'),
        'faultDeletedSuccess': _('تم حذف العطل بنجاح'),
        'errorDeletingFault': _('حدث خطأ في حذف العطل')
    }
    js_translations_json = json.dumps(js_translations, ensure_ascii=False)
    
    # Get all branches for filter
    try:
        branches = Branch.query.all()
    except Exception as e:
        # If branch_type column doesn't exist, use a workaround
        logger.error(f"Error querying branches in camera_faults_sidebar: {e}")
        # Query without the problematic column
        from sqlalchemy import text
        result = db.session.execute(text("SELECT * FROM branch"))
        branches = []
        for row in result:
            branch = Branch()
            branch.id = row[0]
            branch.name = row[1]
            branch.location = row[2]
            branch.ip_address = row[3]
            branch.phone_number = row[4]
            branch.phone_number_2 = row[5]
            branch.phone_number_3 = row[6]
            branch.phone_number_4 = row[7]
            branch.nvr_count = row[8]
            branch.region_id = row[9] if len(row) > 9 else None
            # Set default branch_type if not available
            branch.branch_type = row[10] if len(row) > 10 else 'دائم'
            branches.append(branch)
    technicians = User.query.filter_by(role="technician").all()
    
    # Get all faults with camera and branch info
    all_faults = db.session.query(
        Fault, Camera, Branch
    ).join(
        Camera, Fault.camera_id == Camera.id
    ).join(
        Branch, Camera.branch_id == Branch.id
    ).filter(
        # Filter out BBM transferred faults
        db.or_(
            Fault.repair_notes.is_(None),
            ~Fault.repair_notes.like('%تم النقل إلى قسم BBM%')
        )
    ).order_by(
        Fault.date_reported.desc()
    ).all()
    
    # Convert to format expected by template
    faults = []
    for fault in all_faults:
        faults.append({
            'id': fault[0].id,
            'camera_name': fault[1].name,
            'branch_name': fault[2].name,
            'branch_id': fault[1].branch_id,
            'fault_type': fault[0].fault_type,
            'description': fault[0].description,
            'date_reported': fault[0].date_reported,
            'resolved_at': fault[0].resolved_at,
            'reported_by': fault[0].reported_by,
            'technician_id': fault[0].technician_id,
            'repair_notes': fault[0].repair_notes
        })
    
    return render_template("camera_faults_sidebar.html", 
                         branches=branches, 
                         technicians=technicians,
                         faults=faults,
                         js_translations_json=js_translations_json)

@app.route("/all-camera-faults")
@login_required
def all_camera_faults():
    """Show ALL camera faults without any filters"""
    if current_user.role != "admin":
        flash("فقط المدير يمكنه الوصول لهذه الصفحة ❌", "danger")
        return redirect(url_for("dashboard"))
    
    # Get all branches for filter
    try:
        branches = Branch.query.all()
    except Exception as e:
        logger.error(f"Error querying branches in all_camera_faults: {e}")
        from sqlalchemy import text
        result = db.session.execute(text("SELECT * FROM branch"))
        branches = []
        for row in result:
            branch = Branch()
            branch.id = row[0]
            branch.name = row[1]
            branch.location = row[2]
            branch.ip_address = row[3]
            branch.phone_number = row[4]
            branch.phone_number_2 = row[5]
            branch.phone_number_3 = row[6]
            branch.phone_number_4 = row[7]
            branch.nvr_count = row[8]
            branch.region_id = row[9] if len(row) > 9 else None
            branch.branch_type = row[10] if len(row) > 10 else 'دائم'
            branches.append(branch)
    
    technicians = User.query.filter_by(role="technician").all()
    
    # Get ALL faults with camera and branch info - NO FILTERS
    all_faults = db.session.query(
        Fault, Camera, Branch
    ).join(
        Camera, Fault.camera_id == Camera.id
    ).join(
        Branch, Camera.branch_id == Branch.id
    ).order_by(
        Fault.date_reported.desc()
    ).all()
    
    # Convert to format expected by template
    faults = []
    for fault in all_faults:
        faults.append({
            'id': fault[0].id,
            'camera_name': fault[1].name,
            'branch_name': fault[2].name,
            'branch_id': fault[1].branch_id,
            'fault_type': fault[0].fault_type,
            'description': fault[0].description,
            'date_reported': fault[0].date_reported,
            'resolved_at': fault[0].resolved_at,
            'reported_by': fault[0].reported_by,
            'technician_id': fault[0].technician_id,
            'repair_notes': fault[0].repair_notes
        })
    
    return render_template("camera_faults_sidebar.html", 
                         branches=branches, 
                         technicians=technicians,
                         faults=faults)

@app.route("/api/all-camera-faults")
@login_required
def api_all_camera_faults():
    """API endpoint for ALL camera faults - no filters"""
    # Check if this is an AJAX request
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest' or \
             request.content_type and 'application/json' in request.content_type
    
    if current_user.role != "admin":
        if is_ajax:
            return jsonify({"error": "Unauthorized"}), 403
        return jsonify({"error": "Unauthorized"}), 403
    
    # Get ALL faults with camera and branch info - NO FILTERS
    all_faults = db.session.query(
        Fault, Camera, Branch
    ).join(
        Camera, Fault.camera_id == Camera.id
    ).join(
        Branch, Camera.branch_id == Branch.id
    ).order_by(
        Fault.date_reported.desc()
    ).all()
    
    # Convert to format expected by frontend
    faults = []
    for fault in all_faults:
        faults.append({
            'id': fault[0].id,
            'camera_name': fault[1].name,
            'branch_name': fault[2].name,
            'branch_id': fault[1].branch_id,
            'fault_type': fault[0].fault_type,
            'description': fault[0].description,
            'date_reported': fault[0].date_reported.isoformat() if fault[0].date_reported else None,
            'resolved_at': fault[0].resolved_at.isoformat() if fault[0].resolved_at else None,
            'reported_by': fault[0].reported_by,
            'technician_id': fault[0].technician_id,
            'repair_notes': fault[0].repair_notes
        })
    
    return jsonify({"faults": faults})

@app.route("/api/camera-faults")
@login_required
def api_camera_faults():
    """API endpoint for camera faults - supports fast loading"""
    if current_user.role != "admin":
        return jsonify({"error": "Unauthorized"}), 403
    
    # Get all branches for filter
    branches = Branch.query.all()
    technicians = User.query.filter_by(role="technician").all()
    
    # Get all faults with camera and branch info
    all_faults = db.session.query(
        Fault, Camera, Branch
    ).join(
        Camera, Fault.camera_id == Camera.id
    ).join(
        Branch, Camera.branch_id == Branch.id
    ).filter(
        # Filter out BBM transferred faults
        db.or_(
            Fault.repair_notes.is_(None),
            ~Fault.repair_notes.like('%تم النقل إلى قسم BBM%')
        )
    ).order_by(
        Fault.date_reported.desc()
    ).all()
    
    # Convert to format expected by frontend
    faults = []
    for fault in all_faults:
        # fault is a Row object with 3 elements: Fault, Camera, Branch
        faults.append({
            'id': fault[0].id,  # fault.id
            'camera_name': fault[1].name,  # camera.name
            'branch_name': fault[2].name,  # branch.name
            'branch_id': fault[1].branch_id,  # camera.branch_id
            'fault_type': fault[0].fault_type,  # fault.fault_type
            'description': fault[0].description,  # fault.description
            'date_reported': fault[0].date_reported.isoformat() if fault[0].date_reported else None,
            'resolved_at': fault[0].resolved_at.isoformat() if fault[0].resolved_at else None,
            'reported_by': fault[0].reported_by,  # fault.reported_by
            'technician_id': fault[0].technician_id,  # fault.technician_id
            'repair_notes': fault[0].repair_notes  # fault.repair_notes
        })
    
    return jsonify({"faults": faults})
    camera = Camera.query.get_or_404(camera_id)
    technicians = User.query.filter_by(role='technician').all()
    now = utc_now()

    if current_user.role != "admin":
        flash(_("فقط المدير يمكنه الوصول لهذه الصفحة ❌"), "danger")
        return redirect(url_for("all_faults"))

    # compute average repair duration (in days) for resolved faults of this camera
    faults_list = list(camera.faults)
    resolved = [f for f in faults_list if f.resolved_at]
    if resolved:
        total_seconds = sum((f.resolved_at - f.date_reported).total_seconds() for f in resolved)
        avg_days = total_seconds / len(resolved) / 86400.0
    else:
        avg_days = 0

    return render_template("faults.html", faults=faults_list, camera=camera, technicians=technicians, now=now, avg_days=avg_days)

@app.route("/faults/<int:camera_id>")
@login_required
def faults(camera_id):
    camera = Camera.query.get_or_404(camera_id)
    technicians = User.query.filter_by(role='technician').all()
    now = utc_now()

    if current_user.role != "admin":
        flash(_("فقط المدير يمكنه الوصول لهذه الصفحة ❌"), "danger")
        return redirect(url_for("all_faults"))

    # compute average repair duration (in days) for resolved faults of this camera
    faults_list = list(camera.faults)
    resolved = [f for f in faults_list if f.resolved_at]
    if resolved:
        total_seconds = sum((f.resolved_at - f.date_reported).total_seconds() for f in resolved)
        avg_days = total_seconds / len(resolved) / 86400.0
    else:
        avg_days = 0

    return render_template("faults.html", faults=faults_list, camera=camera, technicians=technicians, now=now, avg_days=avg_days)

@app.route("/convert-to-bbm/<int:fault_id>", methods=["POST"])
@login_required
def convert_to_bbm(fault_id):
    """Convert fault to BBM system and remove from main faults list"""
    if current_user.role != "admin":
        return jsonify({"success": False, "message": "غير مسموح لك بالتحويل إلى BBM ❌"})
    
    fault = Fault.query.get_or_404(fault_id)
    
    try:
        # Create BBM fault record
        bbm_fault = BBMFault(
            original_fault_id=fault.id,
            camera_id=fault.camera_id,
            branch_id=fault.camera.branch_id if fault.camera else None,
            description=fault.description,
            fault_type=fault.fault_type,
            device_type=fault.device_type or "NVR",
            reported_by=fault.reported_by,
            technician_id=fault.technician_id,
            date_reported=fault.date_reported,
            transferred_by=current_user.username,
            status="Pending"
        )
        
        db.session.add(bbm_fault)
        
        # Mark original fault as transferred to BBM and hide from main list
        fault.repair_notes = f"تم التحويل إلى BBM بواسطة {current_user.username} في {utc_now().strftime('%Y-%m-%d %H:%M')} - تم النقل إلى قسم BBM"
        fault.resolved_at = utc_now()
        fault.resolved_by = current_user.username
        fault.status = "محول الي BBM"
        
        db.session.commit()
        
        return jsonify({"success": True, "message": f"تم تحويل العطل #{fault.id} إلى BBM بنجاح ✅"})
        
    except Exception as e:
        print(f"Error converting to BBM: {str(e)}")
        db.session.rollback()
        return jsonify({"success": False, "message": f"حدث خطأ أثناء التحويل إلى BBM: {str(e)} ❌"})

@app.route("/bbm-management")
@login_required
def bbm_management():
    """BBM Management page - show all transferred faults"""
    if current_user.role != "admin":
        flash(_("فقط المدير يمكنه الوصول لصفحة BBM ❌"), "danger")
        return redirect(url_for("dashboard"))
    
    # Get all BBM faults
    bbm_faults = BBMFault.query.order_by(BBMFault.date_transferred.desc()).all()
    
    return render_template("bbm_management.html", 
                         bbm_faults=bbm_faults,
                         branches=Branch.query.all(),
                         technicians=User.query.filter(User.role.in_(["admin", "technician"])).all())


@app.route("/resolve-bbm/<int:bbm_id>", methods=["GET", "POST"])
@login_required
def resolve_bbm(bbm_id):
    """Resolve BBM fault"""
    if current_user.role != "admin":
        flash(_("غير مسموح لك بإصلاح عطل BBM ❌"), "danger")
        return redirect(url_for("dashboard"))
    
    bbm_fault = BBMFault.query.get_or_404(bbm_id)
    
    if request.method == "POST":
        bbm_fault.status = "Resolved"
        bbm_fault.resolved_at = utc_now()
        bbm_fault.resolved_by = current_user.username
        bbm_fault.notes = request.form.get('notes', '')
        
        db.session.commit()
        flash(_("تم إصلاح عطل BBM بنجاح ✅"), "success")
        return redirect(url_for("bbm_management"))
    
    return render_template("resolve_bbm.html", bbm_fault=bbm_fault)


@app.route("/delete-bbm/<int:bbm_id>", methods=["POST"])
@login_required
def delete_bbm(bbm_id):
    """Delete BBM fault"""
    if current_user.role != "admin":
        flash(_("غير مسموح لك بحذف عطل BBM ❌"), "danger")
        return redirect(url_for("dashboard"))
    
    bbm_fault = BBMFault.query.get_or_404(bbm_id)
    
    try:
        db.session.delete(bbm_fault)
        db.session.commit()
        flash(_("تم حذف عطل BBM بنجاح ✅"), "success")
    except Exception as e:
        print(f"Error deleting BBM fault: {str(e)}")
        flash(f"حدث خطأ أثناء حذف عطل BBM: {str(e)} ❌", "danger")
        db.session.rollback()
    
    return redirect(url_for("bbm_management"))

@app.route("/faults/edit/<int:fault_id>", methods=["GET", "POST"])
@login_required
def edit_fault(fault_id):
    fault = Fault.query.get_or_404(fault_id)
    technicians = User.query.filter_by(role='technician').all()
    
    if current_user.role != "admin":
        flash("فقط المدير يمكنه تعديل الأعطال ❌", "danger")
        return redirect(url_for("all_faults"))
    
    if request.method == "POST":
        try:
            # Check if this is an AJAX request
            is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest' or \
                     request.content_type and 'application/json' in request.content_type
            
            # Update all fault fields
            fault.fault_type = request.form.get("fault_type", fault.fault_type)
            fault.device_type = request.form.get("device_type", fault.device_type)
            fault.description = request.form.get("description", fault.description)
            fault.reported_by = request.form.get("reported_by", fault.reported_by)
            
            technician_id = request.form.get("technician_id")
            if technician_id:
                fault.technician_id = int(technician_id)
            else:
                fault.technician_id = None
            
            repair_notes = request.form.get("repair_notes")
            if repair_notes:
                fault.repair_notes = repair_notes
            
            db.session.commit()
            
            if is_ajax:
                return jsonify({
                    'success': True,
                    'message': 'تم تعديل بيانات العطل بنجاح ✅'
                })
            else:
                flash("تم تعديل بيانات العطل بنجاح ✅", "success")
                return redirect(url_for("all_faults"))
            
        except Exception as e:
            db.session.rollback()
            logger.error(f"Error editing fault: {str(e)}", exc_info=True)
            
            if is_ajax:
                return jsonify({
                    'success': False,
                    'message': f'خطأ في تعديل العطل: {str(e)} ❌'
                }), 500
            else:
                flash(f"خطأ في تعديل العطل: {str(e)} ❌", "danger")
    
    return render_template("add_edit_fault.html", fault=fault, camera=fault.camera, technicians=technicians)

@app.route("/faults/resolve/<int:fault_id>", methods=["GET", "POST"])
@login_required
def resolve_fault(fault_id):
    fault = Fault.query.get_or_404(fault_id)

    if current_user.role == "technician" and fault.technician_id != current_user.id:
        flash("غير مسموح لك بإصلاح هذا العطل ❌", "danger")
        return redirect(url_for("all_faults"))

    if request.method == "POST":
        try:
            # Check if this is an AJAX request
            is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest' or \
                     request.content_type and 'application/json' in request.content_type
            
            fault.resolved_at = utc_now()
            fault.resolved_by = current_user.username
            fault.repair_notes = request.form.get('repair_notes', '')

            if 'repair_image' in request.files:
                file = request.files['repair_image']
                if file and file.filename != "" and allowed_upload_file(file.filename, {'png', 'jpg', 'jpeg', 'gif'}):
                    filename = secure_filename(file.filename)
                    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                    file.save(filepath)
                    fault.repair_image = filename
                elif file and file.filename and not allowed_upload_file(file.filename, {'png', 'jpg', 'jpeg', 'gif'}):
                    if is_ajax:
                        return jsonify({
                            'success': False,
                            'message': "صيغة الملف غير مدعومة. استخدم: " + ", ".join(app.config['ALLOWED_IMAGE_EXTENSIONS'])
                        }), 400
                    else:
                        flash("صيغة الملف غير مدعومة. استخدم: " + ", ".join(app.config['ALLOWED_IMAGE_EXTENSIONS']), "warning")

            db.session.commit()
            
            if is_ajax:
                return jsonify({
                    'success': True,
                    'message': 'تم إصلاح العطل بنجاح ✅'
                })
            else:
                flash("تم إصلاح العطل بنجاح  - تم تحديث صفحة جميع الأعطال", "success")
                return redirect(url_for("all_faults"))
                
        except Exception as e:
            db.session.rollback()
            logger.error(f"Error resolving fault: {str(e)}", exc_info=True)
            
            if is_ajax:
                return jsonify({
                    'success': False,
                    'message': f'خطأ في إصلاح العطل: {str(e)} ❌'
                }), 500
            else:
                flash(f"خطأ في إصلاح العطل: {str(e)} ❌", "danger")
                return redirect(url_for("all_faults"))

    return render_template("resolve_fault.html", fault=fault)

@app.route("/faults/delete/<int:fault_id>", methods=["POST"])
@login_required
def delete_fault(fault_id):
    """Delete a fault"""
    if current_user.role != "admin":
        flash("فقط المدير يمكنه حذف الأعطال ❌", "danger")
        return redirect(url_for("dashboard"))
    
    fault = Fault.query.get_or_404(fault_id)
    
    try:
        # Check for any dependencies before deleting
        # Check if this fault has been transferred to BBM
        bbm_fault = BBMFault.query.filter_by(original_fault_id=fault.id).first()
        if bbm_fault:
            error_msg = f"لا يمكن حذف العطل #{fault.id} لأنه مرتبط بأعطال BBM ❌"
            
            # Check if this is an AJAX request
            is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest' or \
                     request.content_type and 'application/json' in request.content_type
            
            if is_ajax:
                return jsonify({"success": False, "message": error_msg})
            else:
                flash(error_msg, "danger")
                return redirect(url_for("all_faults"))
        
        db.session.delete(fault)
        db.session.commit()
        
        success_msg = f"تم حذف العطل #{fault.id} بنجاح ✅"
        
        # Check if this is an AJAX request
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest' or \
                 request.content_type and 'application/json' in request.content_type
        
        if is_ajax:
            return jsonify({"success": True, "message": success_msg})
        else:
            flash(success_msg, "success")
            return redirect(url_for("all_faults"))
        
    except Exception as e:
        db.session.rollback()
        error_msg = str(e)
        
        # Handle specific database errors
        if "FOREIGN KEY constraint" in error_msg:
            error_msg = f"لا يمكن حذف العطل #{fault.id} لأنه مرتبط ببيانات أخرى في النظام ❌"
        elif "no such table" in error_msg:
            error_msg = f"خطأ في قاعدة البيانات: {error_msg} ❌"
        else:
            error_msg = f"حدث خطأ أثناء حذف العطل: {error_msg} ❌"
        
        # Log the error for debugging
        print(f"Error deleting fault {fault_id}: {error_msg}")
        
        # Check if this is an AJAX request
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest' or \
                 request.content_type and 'application/json' in request.content_type
        
        if is_ajax:
            return jsonify({"success": False, "message": error_msg})
        else:
            flash(error_msg, "danger")
            return redirect(url_for("all_faults"))

@app.route("/faults/import-excel", methods=["GET", "POST"])
@login_required
def import_excel_faults():
    """استيراد الأعطال من ملف Excel بالتنسيق المخصص"""
    if current_user.role != "admin":
        flash(_("فقط المدير يمكنه استيراد الأعطال من Excel ❌"), "danger")
        return redirect(url_for("dashboard"))
    
    if request.method == "POST":
        print("=== STARTING EXCEL IMPORT ===")
        
        if "excel_file" not in request.files:
            flash(t('لم يتم اختيار ملف ❌'), "danger")
            return redirect(request.url)
        
        file = request.files["excel_file"]
        
        if file.filename == "":
            flash(t('لم يتم اختيار ملف ❌'), "danger")
            return redirect(request.url)
        
        if not file.filename.endswith((".xlsx", ".xls")):
            flash(t('يجب أن يكون الملف من نوع Excel (.xlsx أو .xls) '), "danger")
            return redirect(request.url)
        
        try:
            import pandas as pd
            df = pd.read_excel(file, header=None, skiprows=1)
            print(f"Excel loaded: {len(df)} rows, {len(df.columns)} columns")
            
            if df.empty:
                flash(t('الملف فارغ أو لا يحتوي على بيانات صالحة'), "danger")
                return redirect(request.url)
            
            # Initialize variables
            faults_added = 0
            errors = []
            seasonal_branches_to_confirm = []
            seasonal_faults_data = []
            
            # Process each row
            for idx, row in df.iterrows():
                row_idx = idx + 2
                
                # Extract data
                camera_name = str(row.iloc[0]).strip() if len(df.columns) >= 1 and pd.notna(row.iloc[0]) else ''
                camera_ip = str(row.iloc[1]).strip() if len(df.columns) >= 2 and pd.notna(row.iloc[1]) else ''
                branch_name = str(row.iloc[2]).strip() if len(df.columns) >= 3 and pd.notna(row.iloc[2]) else ''
                
                print(f"ROW {row_idx}: Camera='{camera_name}', IP='{camera_ip}', Branch='{branch_name}'")
                
                if not camera_name or not branch_name:
                    errors.append(f"Row {row_idx}: Missing camera or branch name")
                    continue
                
                # Find branch
                branch = Branch.query.filter_by(name=branch_name).first()
                if not branch:
                    # Get all available branches for better error message
                    all_branches = Branch.query.with_entities(Branch.name).all()
                    available_branches = [b.name for b in all_branches[:10]]  # Show first 10 branches
                    error_msg = f"Row {row_idx}: Branch '{branch_name}' not found. Available branches: {', '.join(available_branches)}"
                    if len(all_branches) > 10:
                        error_msg += f" (and {len(all_branches) - 10} more)"
                    errors.append(error_msg)
                    print(f"  -> Branch NOT found: '{branch_name}'")
                    print(f"  -> Available branches: {available_branches}")
                    continue
                
                print(f"  -> Branch found: Type='{branch.branch_type}', Region='{branch.region.name if branch.region else 'None'}")
                
                # Find camera
                camera = Camera.query.filter_by(name=camera_name, branch_id=branch.id).first()
                if not camera:
                    errors.append(f"Row {row_idx}: Camera '{camera_name}' not found in branch '{branch_name}'")
                    print(f"  -> Camera NOT found")
                    continue
                
                # Check if seasonal - ONLY North Coast regions are seasonal
                is_seasonal = False
                if branch.region and 'North Coast' in branch.region.name:
                    is_seasonal = True
                    print(f"  -> SEASONAL: North Coast region - {branch.region.name}")
                else:
                    print(f"  -> PERMANENT: Not North Coast region ({branch.region.name if branch.region else 'No region'})")
                
                if is_seasonal:
                    # Add to seasonal list
                    seasonal_fault_data = {
                        'row_idx': row_idx,
                        'camera_name': camera_name,
                        'branch_name': branch_name,
                        'address': branch.location or '',
                        'reported_by': current_user.username,
                        'technician_id': None,
                        'camera_id': camera.id
                    }
                    seasonal_faults_data.append(seasonal_fault_data)
                    if branch.name not in seasonal_branches_to_confirm:
                        seasonal_branches_to_confirm.append(branch.name)
                    print(f"  -> Added to seasonal list. Total: {len(seasonal_faults_data)}")
                else:
                    # Check for existing fault
                    existing_fault = Fault.query.filter_by(
                        camera_id=camera.id,
                        resolved_at=None
                    ).first()
                    
                    if existing_fault:
                        errors.append(f"Row {row_idx}: Camera '{camera_name}' already has an open fault")
                        print(f"  -> Already has open fault")
                        continue
                    
                    # Create fault for permanent branch
                    fault = Fault(
                        camera_id=camera.id,
                        description=f"Camera '{camera_name}' reported offline",
                        fault_type="offline",
                        device_type="Camera",
                        reported_by=current_user.username,
                        technician_id=None
                    )
                    db.session.add(fault)
                    faults_added += 1
                    print(f"  -> Created fault for permanent branch")
            
            print(f"=== SUMMARY ===")
            print(f"Permanent faults added: {faults_added}")
            print(f"Seasonal faults to confirm: {len(seasonal_faults_data)}")
            print(f"Seasonal branches: {seasonal_branches_to_confirm}")
            print(f"Errors: {len(errors)}")
            
            # Save and redirect
            if seasonal_branches_to_confirm:
                print("=== SHOWING CONFIRMATION PAGE ===")
                # Save regular faults first
                if faults_added > 0:
                    db.session.commit()
                    print(f"Committed {faults_added} permanent faults")
                else:
                    db.session.rollback()
                
                # Store seasonal data in session
                session['seasonal_faults_data'] = seasonal_faults_data
                session['seasonal_branches_to_confirm'] = seasonal_branches_to_confirm
                session['regular_faults_added'] = faults_added
                session['excel_errors'] = errors
                
                return render_template("confirm_seasonal_faults.html", 
                                     seasonal_branches=seasonal_branches_to_confirm,
                                     seasonal_faults=seasonal_faults_data,
                                     regular_faults_added=faults_added,
                                     errors=errors)
            else:
                print("=== NO SEASONAL BRANCHES - DIRECT SAVE ===")
                db.session.commit()
                
                if faults_added > 0:
                    flash(f"تم إضافة {faults_added} عطل بنجاح ✅", "success")
                else:
                    flash("لم يتم إضافة أي أعطال", "warning")
                
                if errors:
                    flash(f"هناك {len(errors)} أخطاء: " + " | ".join(errors[:3]), "warning")
                
                return redirect(url_for("all_faults"))
        
        except Exception as e:
            db.session.rollback()
            print(f"ERROR during import: {str(e)}")
            flash(f"حدث خطأ: {str(e)} ❌", "danger")
            return redirect(request.url)
    
    return render_template("import_excel.html")

@app.route("/faults/confirm-seasonal", methods=["POST"])
@login_required
def confirm_seasonal_faults():
    """معالجة تأكيد الفروع الموسمية"""
    if current_user.role != "admin":
        flash("فقط المدير يمكنه تأكيد الأعطال ❌", "danger")
        return redirect(url_for("dashboard"))
    
    confirm = request.form.get('confirm') == 'yes'
    seasonal_faults_data = session.get('seasonal_faults_data', [])
    regular_faults_added = session.get('regular_faults_added', 0)
    errors = session.get('excel_errors', [])
    
    seasonal_faults_added = 0
    
    if confirm and seasonal_faults_data:
        # المستخدم وافق على إضافة أعطال الفروع الموسمية
        try:
            for fault_data in seasonal_faults_data:
                camera = Camera.query.filter_by(name=fault_data['camera_name']).first()
                if camera:
                    fault = Fault(
                        description=f'Camera {fault_data["camera_name"]} is offline - Address: {fault_data["address"]}',
                        fault_type='offline',
                        device_type='Camera',
                        reported_by=fault_data['reported_by'],
                        camera_id=camera.id,
                        technician_id=fault_data['technician_id'],
                        date_reported=utc_now(),
                        expires_at=utc_now() + timedelta(days=7)
                    )
                    db.session.add(fault)
                    seasonal_faults_added += 1
            
            db.session.commit()
            flash(f"✅ تم إضافة {regular_faults_added} عطل للفروع الدائمة و {seasonal_faults_added} عطل للفروع الموسمية بنجاح", "success")
            
        except Exception as e:
            db.session.rollback()
            logger.error(f"خطأ في حفظ أعطال الفروع الموسمية: {str(e)}")
            flash(f"خطأ في حفظ أعطال الفروع الموسمية: {str(e)} ❌", "danger")
    else:
        # المستخدم رفض إضافة أعطال الفروع الموسمية
        flash(f"✅ تم إضافة {regular_faults_added} عطل للفروع الدائمة فقط. تم تجاهل أعطال الفروع الموسمية", "info")
    
    # مسح بيانات الجلسة
    session.pop('seasonal_faults_data', None)
    session.pop('seasonal_branches_to_confirm', None)
    session.pop('regular_faults_added', None)
    session.pop('excel_errors', None)
    
    if errors:
        flash(f"⚠️ هناك {len(errors)} أخطاء أخرى: " + " | ".join(errors[:3]), "warning")
    
    return redirect(url_for("all_faults"))

@app.route("/faults/delete-all", methods=["POST"])
@login_required
def delete_all_faults():
    """حذف جميع الأعطال في النظام"""
    if current_user.role != "admin":
        return jsonify({"success": False, "message": "فقط المدير يمكنه حذف جميع الأعطال ❌"})
    
    try:
        # Check if this is an AJAX request
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest' or \
                 request.content_type and 'application/json' in request.content_type
        
        # Get count before deletion
        total_faults = Fault.query.count()
        
        if total_faults == 0:
            return jsonify({"success": False, "message": "لا توجد أعطال لحذفها"})
        
        # Check for any dependencies before deleting
        bbm_faults_count = BBMFault.query.count()
        if bbm_faults_count > 0:
            error_msg = f"لا يمكن حذف جميع الأعطال لأنه يوجد {bbm_faults_count} عطل مرتبط بأعطال BBM ❌"
            if is_ajax:
                return jsonify({"success": False, "message": error_msg})
            else:
                flash(error_msg, "danger")
                return redirect(url_for("camera_faults_sidebar"))
        
        # Delete all faults
        deleted_count = Fault.query.delete()
        db.session.commit()
        
        success_msg = f"تم حذف {deleted_count} عطل بنجاح ✅"
        
        if is_ajax:
            return jsonify({"success": True, "deleted_count": deleted_count, "message": success_msg})
        else:
            flash(success_msg, "success")
            return redirect(url_for("camera_faults_sidebar"))
        
    except Exception as e:
        db.session.rollback()
        error_msg = f"حدث خطأ أثناء حذف جميع الأعطال: {str(e)} ❌"
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({"success": False, "message": error_msg})
        else:
            flash(error_msg, "danger")
            return redirect(url_for("camera_faults_sidebar"))

@app.route("/faults/download-template")
@login_required
def download_template():
    """تحميل نموذج Excel لاستيراد الأعطال"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    # إنشاء workbook جديد
    wb = Workbook()
    ws = wb.active
    ws.title = "استيراد الأعطال"

    # تعريف الأعمدة
    headers = ["اسم الكاميرا", "نوع العطل", "نوع الجهاز", "الوصف", "المبلّغ", "الفني (اختياري)"]
    
    # تنسيق رؤوس الأعمدة
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # إضافة الرؤوس
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # إضافة بيانات نموذجية (مثال)
    sample_data = [
        ["كاميرا المدخل الرئيسي", "عطل صورة", "NVR", "الصورة غير واضحة ومشوشة", "أحمد محمد", "فني1"],
        ["كاميرا الممر", "عطل تسجيل", "Switch", "لا تسجيل منذ يومين", "محمد علي", "فني2"],
        ["كاميرا الموقف", "عطل اتصال", "Router", "فقدان الاتصال بالشبكة", "سارة حسن", ""],
    ]
    
    # تنسيق بيانات النموذج
    data_alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    data_font = Font(size=11)
    
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = data_alignment
            cell.font = data_font
            cell.border = border

    # ضبط عرض الأعمدة
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15

    # ضبط ارتفاع الصف الأول
    ws.row_dimensions[1].height = 25

    # إضافة ورقة التعليمات
    ws_info = wb.create_sheet("التعليمات")
    ws_info.column_dimensions['A'].width = 40
    ws_info.column_dimensions['B'].width = 50

    # كتابة التعليمات
    info_data = [
        ["", ""],
        ["تعليمات الاستخدام", ""],
        ["", ""],
        ["اسم الكاميرا", "يجب أن يطابق بالضبط أسماء الكاميرات المسجلة في النظام"],
        ["نوع العطل", "أدخل نوع العطل (مثال: عطل صورة، عطل صوت، عطل تسجيل، عطل اتصال)"],
        ["الوصف", "وصف تفصيلي للعطل - بحد أقصى 200 حرف"],
        ["المبلّغ", "اسم الشخص الذي أبلّغ عن العطل"],
        ["الفني (اختياري)", "اسم مستخدم الفني - اترك فارغاً إذا لم تحدد فني"],
        ["", ""],
        ["ملاحظات مهمة:", ""],
        ["1", "تأكد من أن أسماء الكاميرات موجودة في النظام"],
        ["2", "أسماء الفنيين يجب أن تكون أسماء مستخدمين موجودة"],
        ["3", "لا تحذف أو تعيد ترتيب الأعمدة"],
        ["4", "لا تترك أعمدة إلزامية فارغة (الكاميرا والنوع والوصف)"],
        ["5", "صيغة الملف يجب أن تكون .xlsx أو .xls"],
    ]

    info_font = Font(size=11)
    info_title_font = Font(size=12, bold=True, color="FFFFFF")
    info_title_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")

    for row_idx, row_data in enumerate(info_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_info.cell(row=row_idx, column=col_idx, value=value)
            cell.font = info_font
            if row_idx == 2:
                cell.font = info_title_font
                cell.fill = info_title_fill
            if row_idx == 10:
                cell.font = Font(size=11, bold=True)
            cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)

    # حفظ الملف في الذاكرة
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"نموذج_استيراد_الأعطال_{local_now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )

@app.route("/reports/excel")
@login_required
def report_excel():
    """تنزيل تقرير الأعطال بصيغة Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Faults Report"

    headers = [
        "#",
        "Chain",
        "Region",
        "Branch",
        "Camera",
        "Fault Type",
        "Description",
        "Reported By",
        "Reported At",
        "Status",
        "Technician",
        "Repaired At",
        "Repair Notes",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    if current_user.role == "admin":
        faults = Fault.query.order_by(Fault.date_reported.desc()).all()
    else:
        faults = Fault.query.filter(Fault.technician_id == current_user.id).order_by(Fault.date_reported.desc()).all()

    fault_type_map = {
        "كهرباء": "Electrical",
        "شبكة": "Network",
        "جهاز": "Device",
        "برمجيات": "Software",
    }

    for row, f in enumerate(faults, 2):
        region = f.camera.branch.region
        chain_name = (region.chain.name if region and region.chain else "") or ""
        region_name = (region.name if region else "") or ""
        ws.cell(row=row, column=1, value=f.id)
        ws.cell(row=row, column=2, value=chain_name)
        ws.cell(row=row, column=3, value=region_name)
        ws.cell(row=row, column=4, value=f.camera.branch.name)
        ws.cell(row=row, column=5, value=f.camera.name)
        ws.cell(row=row, column=6, value=fault_type_map.get(f.fault_type, f.fault_type or ""))
        ws.cell(row=row, column=7, value=f.description or "")
        ws.cell(row=row, column=8, value=f.reported_by or "")
        ws.cell(row=row, column=9, value=f.date_reported.strftime("%Y-%m-%d %H:%M") if f.date_reported else "")
        ws.cell(row=row, column=10, value="Resolved" if f.resolved_at else "Pending")
        ws.cell(row=row, column=11, value=f.technician.username if f.technician else "")
        ws.cell(row=row, column=12, value=f.resolved_at.strftime("%Y-%m-%d %H:%M") if f.resolved_at else "")
        ws.cell(row=row, column=13, value=(f.repair_notes or "")[:500])

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fn = f"faults_report_{local_now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fn, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# Settings routes (Admin only)
@app.route("/settings", methods=["GET", "POST"])
@login_required
def settings():
    if current_user.role != "admin":
        flash(_("فقط المدير يمكنه الوصول لهذه الصفحة ❌"), "danger")
        return redirect(url_for("dashboard"))
    
    if request.method == "POST":
        # Handle logo upload
        if 'company_logo' in request.files:
            file = request.files['company_logo']
            if file and file.filename != "":
                # Check if file is an image
                allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'svg'}
                if file.filename.rsplit('.', 1)[1].lower() in allowed_extensions:
                    filename = secure_filename(file.filename)
                    filepath = os.path.join(app.config['LOGO_FOLDER'], 'company_logo.png')
                    file.save(filepath)
                    flash("تم رفع لوجو الشركة بنجاح ✅", "success")
        
        # Handle company name
        company_name = request.form.get('company_name', '')
        if company_name:
            app.config['COMPANY_NAME'] = company_name
            flash("تم تحديث اسم الشركة بنجاح ✅", "success")
        
        # Handle logo size
        logo_size = request.form.get('logo_size', '50')
        try:
            logo_size = int(logo_size)
            if 20 <= logo_size <= 200:  # Limit between 20 and 200 pixels
                app.config['LOGO_SIZE'] = logo_size
                flash("تم تحديث حجم اللوجو بنجاح ✅", "success")
            else:
                flash("حجم اللوجو يجب أن يكون بين 20 و 200 بكسل ❌", "danger")
        except ValueError:
            flash("حجم اللوجو غير صحيح ❌", "danger")
        
        return redirect(url_for("settings"))
    
    # Check if logo exists
    logo_exists = os.path.exists(app.config['COMPANY_LOGO'])
    logo_size = app.config.get('LOGO_SIZE', 50)
    
    return render_template("settings.html", 
                          company_name=app.config.get('COMPANY_NAME', 'CCTV Portal EG'),
                          logo_exists=logo_exists,
                          logo_size=logo_size)

import platform
import subprocess
import re

@app.route("/ping-store", methods=["GET", "POST"])
@login_required
def ping_store():
    """
    Ping Store feature - allows users to ping branches by their device IP addresses.
    Search by branch code or name, select Router or NVR device, and ping.
    """
    if current_user.role != "admin":
        flash(_("فقط المدير يمكنه استخدام هذه الميزة"), "danger")
        return redirect(url_for("dashboard"))
    
    ping_result = None
    selected_branch = None
    selected_device = None
    search_query = request.args.get('q', '').strip()
    
    # Get all branches with their devices
    branches_query = Branch.query.join(Region).join(Chain).options(
        db.joinedload(Branch.region).joinedload(Region.chain),
        db.joinedload(Branch.devices)
    )
    
    if search_query:
        # Search by branch name or region name or chain name
        branches_query = branches_query.filter(
            db.or_(
                Branch.name.ilike(f'%{search_query}%'),
                Region.name.ilike(f'%{search_query}%'),
                Chain.name.ilike(f'%{search_query}%')
            )
        )
    
    branches = branches_query.order_by(Branch.name).all()
    
    if request.method == "POST":
        branch_id = request.form.get('branch_id', type=int)
        device_id = request.form.get('device_id', type=int)
        ip_address = request.form.get('ip_address', '').strip()
        
        if not ip_address:
            flash(_("يرجى إدخال عنوان IP"), "danger")
        else:
            # Validate IP address
            ip_pattern = re.compile(r'^(?:(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.){3}(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)$')
            if not ip_pattern.match(ip_address):
                flash(_("عنوان IP غير صحيح"), "danger")
            else:
                # Perform ping
                try:
                    # Test ping command for Windows
                    command = ['ping', '-n', '1', ip_address]
                    
                    print(f"Executing command: {' '.join(command)}")  # Debug line
                    result = subprocess.run(command, capture_output=True, text=True, timeout=10, shell=True)
                    
                    print(f"Return code: {result.returncode}")  # Debug line
                    print(f"Stdout: {result.stdout}")  # Debug line
                    print(f"Stderr: {result.stderr}")  # Debug line
                    
                    ping_result = {
                        'ip': ip_address,
                        'success': result.returncode == 0,
                        'output': result.stdout,
                        'error': result.stderr if result.returncode != 0 else None
                    }
                    
                    if ping_result['success']:
                        flash(_(f"تم الاتصال بنجاح بـ {ip_address}"), "success")
                    else:
                        flash(_(f"فشل الاتصال بـ {ip_address}"), "warning")
                        
                except subprocess.TimeoutExpired:
                    ping_result = {
                        'ip': ip_address,
                        'success': False,
                        'output': None,
                        'error': _('انتهت مهلة الاتصال')
                    }
                    flash(_("انتهت مهلة الاتصال"), "warning")
                except Exception as e:
                    ping_result = {
                        'ip': ip_address,
                        'success': False,
                        'output': None,
                        'error': str(e)
                    }
                    flash(_(f"خطأ في الاتصال: {str(e)}"), "danger")
        
        if branch_id:
            selected_branch = Branch.query.get(branch_id)
        if device_id:
            selected_device = Device.query.get(device_id)
    
    print(f"DEBUG: ping_result = {ping_result}")  # Debug line
    return render_template("ping_store_simple.html",
                           branches=branches,
                           selected_branch=selected_branch,
                           selected_device=selected_device,
                           ping_result=ping_result,
                           search_query=search_query)

@app.route("/upload-document", methods=["GET", "POST"])
@login_required
def upload_document():
    """Upload document and create automatic NVR faults for branch cameras."""
    if request.method == "GET":
        # Get all chains with their regions and branches for the form
        chains = Chain.query.all()
        return render_template("upload_document.html", chains=chains)
    
    if request.method == "POST":
        try:
            # Get form data
            branch_id = request.form.get("branch_id")
            technician_name = request.form.get("technician_name", current_user.username)
            notes = request.form.get("notes", "")
            
            if not branch_id:
                flash(_("يجب اختيار الفرع"), "danger")
                return redirect(request.referrer)
            
            # Get branch and its cameras
            branch = Branch.query.get(branch_id)
            if not branch:
                flash(_("الفرع غير موجود"), "danger")
                return redirect(request.referrer)
            
            # Create automatic NVR fault for all cameras in branch
            cameras = Camera.query.filter_by(branch_id=branch_id).all()
            
            # Get cameras that already have faults before this NVR event
            cameras_with_existing_faults = []
            
            for camera in cameras:
                # Check if there are any existing unresolved faults for this camera (before NVR)
                existing_other_faults = Fault.query.filter(
                    Fault.camera_id == camera.id,
                    Fault.resolved_at.is_(None),
                    Fault.fault_type != 'NVR Issue'
                ).first()
                
                if existing_other_faults:
                    cameras_with_existing_faults.append(camera)
            
            # Delete all existing faults for this branch (except NVR issues and cameras with pre-existing faults)
            all_branch_faults = []
            all_faults = Fault.query.filter(
                Fault.resolved_at.is_(None),
                Fault.fault_type != 'NVR Issue'
            ).all()
            
            # Check each fault's camera branch manually
            for fault in all_faults:
                if fault.camera and fault.camera.branch_id == branch_id:
                    all_branch_faults.append(fault)
            
            # Collect faults to delete (excluding cameras with pre-existing faults)
            faults_to_delete = []
            for fault in all_branch_faults:
                if fault.camera and fault.camera not in cameras_with_existing_faults:
                    faults_to_delete.append(fault)
            
            # Delete the faults
            for fault in faults_to_delete:
                # Delete repair image if exists
                if fault.repair_image:
                    repair_image_path = os.path.join(app.config['UPLOAD_FOLDER'], fault.repair_image)
                    if os.path.exists(repair_image_path):
                        os.remove(repair_image_path)
                db.session.delete(fault)
            
            # Logic for creating NVR faults
            for camera in cameras:
                # Check if there's already an unresolved NVR fault for this camera
                existing_nvr_fault = Fault.query.filter_by(
                    camera_id=camera.id,
                    fault_type='NVR Issue',
                    resolved_at=None
                ).first()
                
                if existing_nvr_fault:
                    continue
                
                # Create NVR fault for this camera
                new_fault = Fault(
                    description=f'NVR issue detected for camera {camera.name}',
                    fault_type='NVR Issue',
                    device_type='NVR',
                    reported_by=technician_name,
                    camera_id=camera.id,
                    repair_notes=notes
                )
                db.session.add(new_fault)
            
            db.session.commit()
            flash(_("تم تسجيل أعطال NVR تلقائياً لجميع كاميرات الفرع"), "success")
            
        except Exception as e:
            flash(_("خطأ في رفع المستند: {}").format(str(e)), "danger")
            db.session.rollback()
    
    return redirect(request.referrer or url_for("dashboard"))

@app.template_filter('sum')
def sum_filter(data, attribute, default=0):
    """Custom filter to sum values from Excel data - skips text, only sums numbers"""
    total = 0
    for item in data:
        try:
            # Get the value from the nested dictionary
            value = item
            for attr in attribute.split('.'):
                if hasattr(value, attr):
                    value = getattr(value, attr)
                elif isinstance(value, dict) and attr in value:
                    value = value[attr]
                else:
                    value = default
                    break
            
            # Skip if value is empty or None
            if value is None or value == '':
                continue
            
            # Convert to number if possible, skip if not a number
            if isinstance(value, str):
                # Try to extract numbers from string, skip if no numbers found
                import re
                numbers = re.findall(r'\d+', str(value))
                if numbers:
                    total += int(numbers[0])  # Take first number found
            elif isinstance(value, (int, float)):
                total += int(value)
        except:
            continue
    return total

@app.template_filter('sum_column')
def sum_column_filter(data, column_name, default=0):
    """Filter to sum values from a specific column in Excel data"""
    total = 0
    for item in data:
        try:
            # Get the row data as dictionary
            row_dict = item.to_dict()
            
            # Get the value for the specific column
            value = row_dict.get(column_name, default)
            
            # Debug: Print what we're trying to sum
            print(f"DEBUG: Column '{column_name}', value: '{value}', type: {type(value)}")
            
            # Skip if value is empty or None
            if value is None or value == '':
                continue
            
            # Convert to number if possible, skip if not a number
            if isinstance(value, str):
                # Try to extract numbers from string, skip if no numbers found
                import re
                # Try different patterns
                numbers = re.findall(r'\d+', str(value))
                if numbers:
                    extracted_num = int(numbers[0])
                    total += extracted_num
                    print(f"DEBUG: Extracted {extracted_num} from '{value}', running total: {total}")
                else:
                    print(f"DEBUG: No numbers found in '{value}'")
            elif isinstance(value, (int, float)):
                total += int(value)
                print(f"DEBUG: Added number {value}, running total: {total}")
        except Exception as e:
            print(f"DEBUG: Error processing '{value}': {e}")
            continue
    
    print(f"DEBUG: Final total for column '{column_name}': {total}")
    return total


@app.route("/total-camera")
@login_required
def total_camera_report():
    """Total Camera page - completely independent Excel data storage"""
    if current_user.role != "admin":
        flash("فقط المدير يمكنه الوصول لهذه الصفحة ❌", "danger")
        return redirect(url_for("dashboard"))
    
    # Get all Excel data records
    excel_data_records = ExcelData.query.order_by(ExcelData.row_number).all()
    
    # Group by filename and get the latest upload
    latest_filename = None
    if excel_data_records:
        latest_filename = excel_data_records[0].excel_filename
    
    # Get data for the latest file
    latest_data = ExcelData.query.filter_by(excel_filename=latest_filename).order_by(ExcelData.row_number).all() if latest_filename else []
    
    return render_template("total_camera_report.html", 
                         excel_data=latest_data, 
                         filename=latest_filename,
                         branches=Branch.query.all())


@app.route("/total-camera/upload-excel", methods=["POST"])
@login_required
def upload_camera_excel():
    """Upload camera data from Excel file - stores raw data"""
    if current_user.role != "admin":
        return jsonify({"error": "Unauthorized"}), 403
    
    try:
        if 'excel_file' not in request.files:
            flash("الرجاء اختيار ملف Excel ❌", "danger")
            return redirect(url_for("total_camera_report"))
        
        file = request.files['excel_file']
        if file.filename == '':
            flash("الرجاء اختيار ملف Excel ❌", "danger")
            return redirect(url_for("total_camera_report"))
        
        if file and allowed_upload_file(file.filename, {'xlsx', 'xls'}):
            filename = secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)
            
            # Read Excel file
            df = pd.read_excel(filepath)
            
            # Debug: Print Excel structure
            print("=== Excel File Structure ===")
            print(f"Columns found: {list(df.columns)}")
            print(f"Rows count: {len(df)}")
            print("First 5 rows:")
            print(df.head().to_string())
            
            # Delete ALL existing Excel data when new file is uploaded
            ExcelData.query.delete()
            db.session.commit()
            
            # Store each row as raw JSON data
            import json
            uploaded_count = 0
            
            for index, row in df.iterrows():
                try:
                    # Convert row to dictionary and store as JSON
                    row_dict = {}
                    for col in df.columns:
                        if pd.notna(row[col]):
                            # Store exactly as it appears in Excel
                            row_dict[str(col)] = str(row[col])
                        else:
                            row_dict[str(col)] = ""
                    
                    # Create ExcelData record
                    excel_record = ExcelData(
                        excel_filename=filename,
                        row_data=json.dumps(row_dict, ensure_ascii=False),
                        row_number=index + 1
                    )
                    db.session.add(excel_record)
                    uploaded_count += 1
                    
                except Exception as e:
                    print(f"Error processing row {index}: {str(e)}")
                    continue
            
            db.session.commit()
            
            flash(f"✅ تم رفع {uploaded_count} صف بنجاح من ملف {filename}!", "success")
            return redirect(url_for("total_camera_report"))
        
        else:
            flash("صيغة الملف غير مدعومة. استخدم ملف Excel (.xlsx أو .xls) ❌", "danger")
            return redirect(url_for("total_camera_report"))
            
    except Exception as e:
        print(f"Upload error: {str(e)}")
        flash(f"خطأ في رفع الملف: {str(e)} ❌", "danger")
        return redirect(url_for("total_camera_report"))


@app.route("/total-camera/save-changes", methods=["POST"])
@login_required
def save_camera_changes():
    """Save edited changes to Excel data"""
    if current_user.role != "admin":
        return jsonify({"error": "Unauthorized"}), 403
    
    try:
        data = request.get_json()
        if not data or 'data' not in data:
            return jsonify({"success": False, "error": "No data provided"}), 400
        
        excel_data = data['data']
        filename = f"edited_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Delete all existing Excel data
        ExcelData.query.delete()
        db.session.commit()
        
        # Store new data
        import json
        uploaded_count = 0
        
        for index, row_data in enumerate(excel_data):
            try:
                # Create ExcelData record
                excel_record = ExcelData(
                    excel_filename=filename,
                    row_data=json.dumps(row_data, ensure_ascii=False),
                    row_number=index + 1
                )
                db.session.add(excel_record)
                uploaded_count += 1
                
            except Exception as e:
                print(f"Error processing row {index}: {str(e)}")
                continue
        
        db.session.commit()
        
        return jsonify({
            "success": True, 
            "message": f"Successfully saved {uploaded_count} rows",
            "filename": filename
        })
        
    except Exception as e:
        print(f"Save error: {str(e)}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/total-camera/download-excel")
@login_required
def download_camera_excel():
    """Download Excel data as Excel file"""
    if current_user.role != "admin":
        flash("فقط المدير يمكنه تنزيل البيانات ❌", "danger")
        return redirect(url_for("dashboard"))
    
    try:
        print("Starting Excel download process...")
        
        # Get latest Excel data
        excel_data_records = ExcelData.query.order_by(ExcelData.upload_date.desc()).first()
        if not excel_data_records:
            flash("لا توجد بيانات للتنزيل ❌", "warning")
            return redirect(url_for("total_camera_report"))
        
        print(f"Found Excel data record: {excel_data_records.excel_filename}")
        
        # Get all records for the latest file
        latest_filename = excel_data_records.excel_filename
        excel_records = ExcelData.query.filter_by(excel_filename=latest_filename).order_by(ExcelData.row_number).all()
        
        print(f"Found {len(excel_records)} records for {latest_filename}")
        
        if not excel_records:
            flash("لا توجد بيانات للتنزيل ❌", "warning")
            return redirect(url_for("total_camera_report"))
        
        # Convert back to DataFrame
        import json
        data_list = []
        for record in excel_records:
            row_dict = record.to_dict()
            data_list.append(row_dict)
        
        print(f"Converted {len(data_list)} records to list")
        
        df = pd.DataFrame(data_list)
        print(f"Created DataFrame with shape: {df.shape}")
        
        # Create Excel file
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = 'Camera Data'
        
        # Add headers (from first row keys)
        if data_list:
            headers = list(data_list[0].keys())
            ws.append(headers)
            print(f"Added headers: {headers}")
            
            # Add data
            for i, row_data in enumerate(data_list):
                row_values = [row_data.get(header, "") for header in headers]
                ws.append(row_values)
                if i < 5:  # Log first 5 rows
                    print(f"Row {i}: {row_values}")
        
        print("Saving workbook...")
        # Save to memory
        wb.save(output)
        output.seek(0)
        
        file_size = len(output.getvalue())
        print(f"Workbook saved, size: {file_size} bytes")
        
        # Check if file is too large
        if file_size > 50 * 1024 * 1024:  # 50MB limit
            flash("الملف كبير جداً للتنزيل ❌", "warning")
            return redirect(url_for("total_camera_report"))
        
        # Create response with proper headers
        response = make_response(output.getvalue())
        response.headers['Content-Disposition'] = f'attachment; filename="{latest_filename}"'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Length'] = str(file_size)
        
        print(f"Created response with headers: {dict(response.headers)}")
        return response
        
    except Exception as e:
        print(f"Download error: {str(e)}")
        import traceback
        traceback.print_exc()
        flash(f"خطأ في تنزيل البيانات: {str(e)} ❌", "danger")
        return redirect(url_for("total_camera_report"))


@app.route("/total-camera/delete-all-data", methods=["POST"])
@login_required
def delete_all_camera_data():
    """Delete all Excel data from the total camera page"""
    if current_user.role != "admin":
        flash(_("فقط المدير يمكنه حذف البيانات ❌"), "danger")
        return redirect(url_for("dashboard"))
    
    try:
        # Delete all ExcelData records
        deleted_count = ExcelData.query.count()
        ExcelData.query.delete()
        db.session.commit()
        
        flash(f"تم حذف جميع البيانات بنجاح! ({deleted_count} سجل تم حذفه) ✅", "success")
        print(f"Admin {current_user.username} deleted all camera data ({deleted_count} records)")
        
    except Exception as e:
        db.session.rollback()
        print(f"Error deleting all camera data: {str(e)}")
        flash(f"حدث خطأ أثناء حذف البيانات: {str(e)} ❌", "danger")
    
    return redirect(url_for("total_camera_report"))


@app.route("/enhanced-reports")
@login_required
def enhanced_reports():
    """Enhanced reports page with branch-wise fault summary and Excel data"""
    if current_user.role != "admin":
        flash(_("فقط المدير يمكنه الوصول إلى التقارير المفصلة ❌"), "danger")
        return redirect(url_for("dashboard"))
    
    # Get search parameters
    search_branch = request.args.get('search_branch', '').strip()
    
    # Get all branches with fault statistics
    try:
        query = Branch.query
        if search_branch:
            # Search by branch number (ID) or name
            if search_branch.isdigit():
                query = query.filter(Branch.id == int(search_branch))
            else:
                query = query.filter(Branch.name.contains(search_branch))
        
        branches = query.all()
    except Exception as e:
        # If branch_type column doesn't exist, use a workaround
        logger.error(f"Error querying branches: {e}")
        # Query without the problematic column
        from sqlalchemy import text
        result = db.session.execute(text("SELECT * FROM branch"))
        branches = []
        for row in result:
            branch = Branch()
            branch.id = row[0]
            branch.name = row[1]
            branch.location = row[2]
            branch.ip_address = row[3]
            branch.phone_number = row[4]
            branch.phone_number_2 = row[5]
            branch.phone_number_3 = row[6]
            branch.phone_number_4 = row[7]
            branch.nvr_count = row[8]
            branch.region_id = row[9] if len(row) > 9 else None
            # Set default branch_type if not available
            branch.branch_type = row[10] if len(row) > 10 else 'دائم'
            branches.append(branch)
    
    for branch in branches:
        # Count faults for this branch
        camera_faults = Fault.query.join(Camera).filter(Camera.branch_id == branch.id).all()
        branch.total_faults = len(camera_faults)
        branch.open_faults = len([f for f in camera_faults if not f.resolved_at])
        branch.closed_faults = len([f for f in camera_faults if f.resolved_at])
    
    # Get Excel data
    excel_data_records = ExcelData.query.order_by(ExcelData.upload_date.desc()).first()
    excel_data = []
    filename = None
    
    if excel_data_records:
        filename = excel_data_records.excel_filename
        excel_data = ExcelData.query.filter_by(excel_filename=filename).order_by(ExcelData.row_number).all()
        excel_data = [record.to_dict() for record in excel_data]
    
    return render_template("enhanced_reports.html", 
                         branches=branches, 
                         excel_data=excel_data,
                         filename=filename)

@app.route("/branch-faults/<int:branch_id>")
@login_required
def branch_faults(branch_id):
    """Show faults for specific branch"""
    if current_user.role != "admin":
        flash("فقط المدير يمكنه مشاهدة تفاصيل الفروع ❌", "danger")
        return redirect(url_for("dashboard"))
    
    branch = Branch.query.get_or_404(branch_id)
    faults = Fault.query.join(Camera).filter(Camera.branch_id == branch_id).order_by(Fault.date_reported.desc()).all()
    
    return render_template("branch_faults.html", 
                         branch=branch, 
                         faults=faults)

@app.route("/branch-faults-excel/<int:branch_id>")
@login_required
def branch_faults_excel(branch_id):
    """Download Excel report for specific branch"""
    if current_user.role != "admin":
        flash("فقط المدير يمكنه تنزيل تقارير الفروع ❌", "danger")
        return redirect(url_for("dashboard"))
    
    branch = Branch.query.get_or_404(branch_id)
    faults = Fault.query.join(Camera).filter(Camera.branch_id == branch_id).order_by(Fault.date_reported.desc()).all()
    
    try:
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f'Branch {branch.name} Faults Report'
        
        # Define styles
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # Write headers
        headers = [
            'ID', 'Camera Name', 'Fault Type', 'Description', 'Reported By', 
            'Date Reported', 'Status', 'Resolved By', 'Date Resolved', 'Repair Notes'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data
        row_num = 2
        for fault in faults:
            ws.cell(row=row_num, column=1, value=fault.id)
            ws.cell(row=row_num, column=2, value=fault.camera.name if fault.camera else '')
            ws.cell(row=row_num, column=3, value=fault.fault_type)
            ws.cell(row=row_num, column=4, value=fault.description)
            ws.cell(row=row_num, column=5, value=fault.reported_by)
            ws.cell(row=row_num, column=6, value=fault.date_reported.strftime('%Y-%m-%d %H:%M:%S') if fault.date_reported else '')
            ws.cell(row=row_num, column=7, value='Resolved' if fault.resolved_at else 'Open')
            ws.cell(row=row_num, column=8, value=fault.resolved_by or '')
            ws.cell(row=row_num, column=9, value=fault.resolved_at.strftime('%Y-%m-%d %H:%M:%S') if fault.resolved_at else '')
            ws.cell(row=row_num, column=10, value=fault.repair_notes or '')
            row_num += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to memory
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return send_file(
            excel_file,
            as_attachment=True,
            download_name=f"branch_{branch.name}_faults_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f"Error generating branch faults report: {str(e)}", "danger")
        return redirect(url_for("enhanced_reports"))

@app.route("/reports")
@login_required
def reports():
    """Reports page - main dashboard for all reports."""
    if current_user.role != "admin":
        flash("فقط المدير يمكنه الوصول إلى التقارير ❌", "danger")
        return redirect(url_for("dashboard"))
    
    # Get statistics
    camera_faults_count = Fault.query.count()
    device_faults_count = DeviceFault.query.count()
    total_faults_count = camera_faults_count + device_faults_count
    resolved_faults_count = Fault.query.filter(Fault.resolved_at.isnot(None)).count() + \
                           DeviceFault.query.filter(DeviceFault.resolved_at.isnot(None)).count()
    
    return render_template("reports.html",
                           camera_faults_count=camera_faults_count,
                           device_faults_count=device_faults_count,
                           total_faults_count=total_faults_count,
                           resolved_faults_count=resolved_faults_count)

@app.route("/reports/camera-repaired")
@login_required
def camera_repaired_report():
    """Generate and download repaired cameras report in Excel format."""
    from datetime import datetime
    try:
        # Get all repaired camera faults
        faults = Fault.query.filter(Fault.resolved_at.isnot(None)).join(Camera).all()
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Repaired Cameras Report"
        
        # Define styles
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='28a745', end_color='28a745', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # Write headers
        headers = [
            'ID', 'Camera Name', 'Branch', 'Region', 'Chain',
            'Fault Type', 'Description', 'Reported By', 'Date Reported',
            'Status', 'Resolved By', 'Date Resolved', 'Repair Notes'
        ]
        
        # Apply header styles
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data
        row_num = 2
        for fault in faults:
            ws.cell(row=row_num, column=1, value=fault.id)
            ws.cell(row=row_num, column=2, value=fault.camera.name if fault.camera else '')
            ws.cell(row=row_num, column=3, value=fault.camera.branch.name if fault.camera and fault.camera.branch else '')
            ws.cell(row=row_num, column=4, value=fault.camera.branch.region.name if fault.camera and fault.camera.branch and fault.camera.branch.region else '')
            ws.cell(row=row_num, column=5, value=fault.camera.branch.region.chain.name if fault.camera and fault.camera.branch and fault.camera.branch.region and fault.camera.branch.region.chain else '')
            ws.cell(row=row_num, column=6, value=fault.fault_type)
            ws.cell(row=row_num, column=7, value=fault.description)
            ws.cell(row=row_num, column=8, value=fault.reported_by)
            ws.cell(row=row_num, column=9, value=fault.date_reported.strftime('%Y-%m-%d %H:%M') if fault.date_reported else '')
            ws.cell(row=row_num, column=10, value='Resolved' if fault.resolved_at else 'Open')
            ws.cell(row=row_num, column=11, value=fault.resolved_by or '')
            ws.cell(row=row_num, column=12, value=fault.resolved_at.strftime('%Y-%m-%d %H:%M') if fault.resolved_at else '')
            ws.cell(row=row_num, column=13, value=fault.repair_notes or '')
            row_num += 1
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Save to memory
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return send_file(
            excel_file,
            as_attachment=True,
            download_name=f"repaired_cameras_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f"Error generating repaired cameras report: {str(e)}", "danger")
        return redirect(url_for("reports"))

@app.route("/reports/camera-faults")
@login_required
def camera_faults_report():
    """Generate and download camera faults report in Excel format."""
    from datetime import datetime
    try:
        # Get all cameras with offline status and create automatic faults
        offline_cameras = Camera.query.filter_by(status='offline').all()
        
        # Create automatic faults for offline cameras if they don't exist
        for camera in offline_cameras:
            # Check if there's an NVR fault for this camera's branch
            branch_nvr_faults = []
            all_nvr_faults = Fault.query.filter(
                Fault.fault_type == 'NVR Issue',
                Fault.resolved_at.is_(None)
            ).all()
            
            # Check each NVR fault's camera branch manually
            for nvr_fault in all_nvr_faults:
                if nvr_fault.camera and nvr_fault.camera.branch_id == camera.branch_id:
                    branch_nvr_faults.append(nvr_fault)
            
            # If branch has NVR faults, don't create new offline faults
            if branch_nvr_faults:
                continue
            
            existing_fault = Fault.query.filter_by(
                camera_id=camera.id,
                fault_type='offline',
                resolved_at=None
            ).first()
            
            if not existing_fault:
                # Create automatic fault for offline camera
                new_fault = Fault(
                    description=f'Camera {camera.name} is offline',
                    fault_type='offline',
                    device_type='Camera',
                    reported_by='System',
                    camera_id=camera.id
                )
                db.session.add(new_fault)
        
        db.session.commit()
        
        # Get all unresolved camera faults (including offline cameras)
        faults = Fault.query.filter(Fault.resolved_at.is_(None)).join(Camera).all()
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Camera Faults Report"
        
        # Define styles
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # Write headers
        headers = [
            'ID', 'Camera Name', 'Branch', 'Region', 'Chain',
            'Fault Type', 'Description', 'Reported By', 'Date Reported',
            'Status', 'Priority', 'Days Open'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data
        for row_num, fault in enumerate(faults, 2):
            # Calculate days open
            days_open = 0
            if fault.date_reported:
                from datetime import datetime
                days_open = (datetime.now() - fault.date_reported).days
            
            ws.cell(row=row_num, column=1, value=fault.id)
            ws.cell(row=row_num, column=2, value=fault.camera.name if fault.camera else 'N/A')
            ws.cell(row=row_num, column=3, value=fault.camera.branch.name if fault.camera and fault.camera.branch else 'N/A')
            ws.cell(row=row_num, column=4, value=fault.camera.branch.region.name if fault.camera and fault.camera.branch and fault.camera.branch.region else 'N/A')
            ws.cell(row=row_num, column=5, value=fault.camera.branch.region.chain.name if fault.camera and fault.camera.branch and fault.camera.branch.region and fault.camera.branch.region.chain else 'N/A')
            ws.cell(row=row_num, column=6, value=fault.fault_type)
            ws.cell(row=row_num, column=7, value=fault.description)
            ws.cell(row=row_num, column=8, value=fault.reported_by)
            ws.cell(row=row_num, column=9, value=fault.date_reported.strftime('%Y-%m-%d %H:%M:%S') if fault.date_reported else '')
            ws.cell(row=row_num, column=10, value='Open' if not fault.resolved_at else 'Resolved')
            ws.cell(row=row_num, column=11, value='Medium')
            ws.cell(row=row_num, column=12, value=days_open)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to memory
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Create response
        response = make_response(excel_file.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=camera_faults_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return response
        
    except Exception as e:
        flash(_(f"خطأ في إنشاء التقرير: {str(e)}"), "danger")
        return redirect(url_for("reports"))

@app.route("/reports/device-repaired")
@login_required
def device_repaired_report():
    """Generate and download repaired devices report in Excel format."""
    from datetime import datetime
    try:
        # Get all repaired device faults
        faults = DeviceFault.query.filter(DeviceFault.resolved_at.isnot(None)).join(Device).all()
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Repaired Devices Report"
        
        # Define styles
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='17a2b8', end_color='17a2b8', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # Write headers
        headers = [
            'ID', 'Device Name', 'Device Type', 'Branch', 'Region', 'Chain',
            'Fault Type', 'Description', 'Reported By', 'Date Reported',
            'Status', 'Resolved By', 'Date Resolved', 'Repair Notes'
        ]
        
        # Apply header styles
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data
        row_num = 2
        for fault in faults:
            ws.cell(row=row_num, column=1, value=fault.id)
            ws.cell(row=row_num, column=2, value=fault.device.name if fault.device else '')
            ws.cell(row=row_num, column=3, value=fault.device.device_type if fault.device else '')
            ws.cell(row=row_num, column=4, value=fault.device.branch.name if fault.device and fault.device.branch else '')
            ws.cell(row=row_num, column=5, value=fault.device.branch.region.name if fault.device and fault.device.branch and fault.device.branch.region else '')
            ws.cell(row=row_num, column=6, value=fault.device.branch.region.chain.name if fault.device and fault.device.branch and fault.device.branch.region and fault.device.branch.region.chain else '')
            ws.cell(row=row_num, column=7, value=fault.fault_type)
            ws.cell(row=row_num, column=8, value=fault.description)
            ws.cell(row=row_num, column=9, value=fault.reported_by)
            ws.cell(row=row_num, column=10, value=fault.date_reported.strftime('%Y-%m-%d %H:%M') if fault.date_reported else '')
            ws.cell(row=row_num, column=11, value='Resolved' if fault.resolved_at else 'Open')
            ws.cell(row=row_num, column=12, value=fault.resolved_by or '')
            ws.cell(row=row_num, column=13, value=fault.resolved_at.strftime('%Y-%m-%d %H:%M') if fault.resolved_at else '')
            ws.cell(row=row_num, column=14, value=fault.repair_notes or '')
            row_num += 1
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Save to memory
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return send_file(
            excel_file,
            as_attachment=True,
            download_name=f"repaired_devices_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f"Error generating repaired devices report: {str(e)}", "danger")
        return redirect(url_for("reports"))

@app.route("/reports/device-faults")
@login_required
def device_faults_report():
    """Generate and download device faults report in Excel format."""
    from datetime import datetime
    try:
        # Get all unresolved device faults only
        faults = DeviceFault.query.filter(DeviceFault.resolved_at.is_(None)).join(Device).all()
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Device Faults Report"
        
        # Define styles
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='28a745', end_color='28a745', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # Write headers
        headers = [
            'ID', 'Device Name', 'Device Type', 'Branch', 'Region', 'Chain',
            'Fault Type', 'Description', 'Reported By', 'Date Reported',
            'Status', 'Priority', 'Days Open'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data
        for row_num, fault in enumerate(faults, 2):
            ws.cell(row=row_num, column=1, value=fault.id)
            ws.cell(row=row_num, column=2, value=fault.device.name if fault.device else 'N/A')
            ws.cell(row=row_num, column=3, value=fault.device.device_type if fault.device else 'N/A')
            ws.cell(row=row_num, column=4, value=fault.device.branch.name if fault.device and fault.device.branch else 'N/A')
            ws.cell(row=row_num, column=5, value=fault.device.branch.region.name if fault.device and fault.device.branch and fault.device.branch.region else 'N/A')
            ws.cell(row=row_num, column=6, value=fault.device.branch.region.chain.name if fault.device and fault.device.branch and fault.device.branch.region and fault.device.branch.region.chain else 'N/A')
            ws.cell(row=row_num, column=7, value=fault.fault_type)
            ws.cell(row=row_num, column=8, value=fault.description)
            ws.cell(row=row_num, column=9, value=fault.reported_by)
            ws.cell(row=row_num, column=10, value=fault.date_reported.strftime('%Y-%m-%d %H:%M:%S') if fault.date_reported else '')
            ws.cell(row=row_num, column=11, value='Resolved' if fault.resolved_at else 'Pending')
            ws.cell(row=row_num, column=12, value=fault.technician.username if fault.technician else 'Not Specified')
            ws.cell(row=row_num, column=13, value=fault.resolved_by or '')
            ws.cell(row=row_num, column=14, value=fault.resolved_at.strftime('%Y-%m-%d %H:%M:%S') if fault.resolved_at else '')
            ws.cell(row=row_num, column=15, value=fault.repair_notes or '')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to memory
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Create response
        response = make_response(excel_file.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=device_faults_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return response
        
    except Exception as e:
        flash(_(f"خطأ في إنشاء التقرير: {str(e)}"), "danger")
        return redirect(url_for("reports"))

@app.route("/reports/all-faults")
@login_required
def all_faults_report():
    """Generate and download all faults report in Excel format."""
    try:
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "All Faults Report"
        
        # Define styles
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='ffc107', end_color='ffc107', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # Write headers
        headers = [
            'Report Type', 'ID', 'Item Name', 'Item Type', 'Branch', 'Region', 'Chain',
            'Fault Type', 'Description', 'Reported By', 'Date Reported',
            'Status', 'Technician/Resolved By', 'Date Resolved', 'Repair Notes'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write camera faults
        camera_faults = Fault.query.all()
        for fault in camera_faults:
            ws.append([
                'Camera Fault',
                fault.id,
                fault.camera.name if fault.camera else 'N/A',
                'Camera',
                fault.camera.branch.name if fault.camera and fault.camera.branch else 'N/A',
                fault.camera.branch.region.name if fault.camera and fault.camera.branch and fault.camera.branch.region else 'N/A',
                fault.camera.branch.region.chain.name if fault.camera and fault.camera.branch and fault.camera.branch.region and fault.camera.branch.region.chain else 'N/A',
                fault.fault_type,
                fault.description,
                fault.reported_by,
                fault.date_reported.strftime('%Y-%m-%d %H:%M:%S') if fault.date_reported else '',
                'Resolved' if fault.resolved_at else 'Pending',
                fault.resolved_by or '',
                fault.resolved_at.strftime('%Y-%m-%d %H:%M:%S') if fault.resolved_at else '',
                fault.repair_notes or ''
            ])
        
        # Write device faults
        device_faults = DeviceFault.query.all()
        for fault in device_faults:
            ws.append([
                'Device Fault',
                fault.id,
                fault.device.name if fault.device else 'N/A',
                fault.device.device_type if fault.device else 'N/A',
                fault.device.branch.name if fault.device and fault.device.branch else 'N/A',
                fault.device.branch.region.name if fault.device and fault.device.branch and fault.device.branch.region else 'N/A',
                fault.device.branch.region.chain.name if fault.device and fault.device.branch and fault.device.branch.region and fault.device.branch.region.chain else 'N/A',
                fault.fault_type,
                fault.description,
                fault.reported_by,
                fault.date_reported.strftime('%Y-%m-%d %H:%M:%S') if fault.date_reported else '',
                'Resolved' if fault.resolved_at else 'Pending',
                fault.technician.username if fault.technician else 'Not Specified',
                fault.resolved_by or '',
                fault.resolved_at.strftime('%Y-%m-%d %H:%M:%S') if fault.resolved_at else '',
                fault.repair_notes or ''
            ])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to memory
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Create response
        response = make_response(excel_file.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=all_faults_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return response
        
    except Exception as e:
        flash(_(f"خطأ في إنشاء التقرير: {str(e)}"), "danger")
        return redirect(url_for("reports"))

@app.route("/reports/branch-closure")
@login_required
def branch_closure_report():
    """Generate and download branch open/close history report in Excel format."""
    try:
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Branch Open Close History"
        
        # Define styles
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='6c757d', end_color='6c757d', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # Write headers
        headers = [
            'Branch ID', 'Branch Name', 'Location', 'Region', 'Chain',
            'Action', 'Action Date', 'Reason', 'Reporter Name',
            'Phone Number', 'IP Address', 'Branch Type'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Get all branch history records
        history_records = BranchHistory.query.order_by(BranchHistory.action_date.desc()).all()
        
        for record in history_records:
            # Format action for display
            action_display = 'Closed' if record.action == 'close' else 'Opened'
            
            ws.append([
                record.branch_id,
                record.branch.name if record.branch else 'N/A',
                record.branch.location if record.branch else '',
                record.branch.region.name if record.branch and record.branch.region else '',
                record.branch.region.chain.name if record.branch and record.branch.region and record.branch.region.chain else '',
                action_display,
                record.action_date.strftime('%Y-%m-%d %H:%M:%S') if record.action_date else '',
                record.reason or '',
                record.reporter_name or '',
                record.branch.phone_number if record.branch else '',
                record.branch.ip_address if record.branch else '',
                record.branch.branch_type if record.branch else ''
            ])
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to memory
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Create response
        response = make_response(excel_file.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=branch_open_close_history_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return response
        
    except Exception as e:
        flash(_(f"خطأ في إنشاء تقرير سجل فتح وإغلاق الفروع: {str(e)}"), "danger")
        return redirect(url_for("reports"))

@app.route("/quick-ping", methods=["POST"])
@login_required
def quick_ping():
    """Quick ping API endpoint for AJAX requests."""
    # Check if this is an AJAX request
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest' or \
             request.content_type and 'application/json' in request.content_type
    
    if current_user.role != "admin":
        if is_ajax:
            return jsonify({'error': 'Unauthorized', 'success': False}), 403
        return jsonify({'error': 'Unauthorized'}), 403
    
    try:
        data = request.get_json()
        ip_address = data.get('ip_address', '').strip()
        
        if not ip_address:
            return jsonify({'error': 'IP address is required', 'success': False}), 400
        
        # Validate IP address
        ip_pattern = re.compile(r'^(?:(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.){3}(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)$')
        if not ip_pattern.match(ip_address):
            return jsonify({'error': 'Invalid IP address', 'success': False}), 400
        
        # Perform ping
        command = ['ping', '-n', '1', ip_address]
        result = subprocess.run(command, capture_output=True, text=True, timeout=10, shell=True)
        
        ping_result = {
            'ip': ip_address,
            'success': result.returncode == 0,
            'output': result.stdout,
            'error': result.stderr if result.returncode != 0 else None
        }
        
        return jsonify(ping_result)
        
    except subprocess.TimeoutExpired:
        return jsonify({
            'ip': ip_address if 'ip_address' in locals() else 'unknown',
            'success': False,
            'output': None,
            'error': 'Connection timeout'
        })
    except Exception as e:
        return jsonify({
            'ip': ip_address if 'ip_address' in locals() else 'unknown',
            'success': False,
            'output': None,
            'error': str(e)
        }), 500

@app.route("/api/branch-devices/<int:branch_id>")
@login_required
def get_branch_devices(branch_id):
    """API endpoint to get devices for a specific branch."""
    if current_user.role != "admin":
        return jsonify({'error': 'Unauthorized'}), 403
    
    branch = Branch.query.get_or_404(branch_id)
    devices = Device.query.filter_by(branch_id=branch_id).all()
    
    devices_data = [{
        'id': d.id,
    } for d in devices if d.ip_address]
    
    return jsonify({
        'branch_id': branch_id,
        'branch_name': branch.name,
        'devices': devices_data
    })

@app.route("/api/all-camera-faults/download-excel")
@login_required
def download_filtered_faults_excel():
    """تنزيل الأعطال المفلترة كـ Excel - بناءً على الفلاتر المطبقة في camera-faults-sidebar"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill

        # الحصول على الفلاتر من query parameters
        search_term = request.args.get('search', '').strip()
        status_filter = request.args.get('status', 'all')
        branch_filter = request.args.get('branch', 'all')

        # بناء query الأساسي
        query = Fault.query.join(Camera).join(Branch)

        # تطبيق الفلاتر
        if search_term:
            query = query.filter(
                db.or_(
                    Camera.name.ilike(f'%{search_term}%'),
                    Fault.description.ilike(f'%{search_term}%'),
                    Branch.name.ilike(f'%{search_term}%')
                )
            )

        if status_filter == 'open':
            query = query.filter(Fault.resolved_at.is_(None))
        elif status_filter == 'resolved':
            query = query.filter(Fault.resolved_at.isnot(None))

        if branch_filter != 'all':
            query = query.filter(Branch.id == int(branch_filter))

        # تنفيذ الاستعلام وترتيب النتائج
        faults = query.order_by(Fault.date_reported.desc()).all()

        # إنشاء ملف Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Camera Faults Report"

        # تنسيق الرؤوس
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        # رؤوس الأعمدة
        headers = [
            '#',
            'اسم الكاميرا',
            'الفرع',
            'نوع العطل',
            'الوصف',
            'الحالة',
            'تاريخ الإبلاغ',
            'المبلغ',
            'الفني المسؤول',
            'تاريخ الإصلاح',
            'ملاحظات الإصلاح'
        ]

        # إضافة الرؤوس
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # إضافة البيانات
        for row_num, fault in enumerate(faults, 2):
            # تحديد الحالة
            if fault.repair_notes and ('تم النقل إلى قسم BBM' in fault.repair_notes or 'Fault successfully converted to BBM' in fault.repair_notes):
                status = 'محول إلى BBM'
            elif fault.resolved_at:
                status = 'مُصلح'
            else:
                status = 'مفتوح'

            ws.cell(row=row_num, column=1, value=fault.id)
            ws.cell(row=row_num, column=2, value=fault.camera.name if fault.camera else 'غير محدد')
            ws.cell(row=row_num, column=3, value=fault.camera.branch.name if fault.camera and fault.camera.branch else 'غير محدد')
            ws.cell(row=row_num, column=4, value=fault.fault_type)
            ws.cell(row=row_num, column=5, value=fault.description)
            ws.cell(row=row_num, column=6, value=status)
            ws.cell(row=row_num, column=7, value=fault.date_reported.strftime('%Y-%m-%d %H:%M') if fault.date_reported else '')
            ws.cell(row=row_num, column=8, value=fault.reported_by or '')
            ws.cell(row=row_num, column=9, value=fault.technician.username if fault.technician else 'غير محدد')
            ws.cell(row=row_num, column=10, value=fault.resolved_at.strftime('%Y-%m-%d %H:%M') if fault.resolved_at else '')
            ws.cell(row=row_num, column=11, value=fault.repair_notes or '')

        # تعديل عرض الأعمدة تلقائياً
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # حفظ الملف في الذاكرة
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # إنشاء اسم الملف مع التاريخ والوقت
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"أعطال_الكاميرات_{timestamp}.xlsx"

        return send_file(
            excel_file,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        flash(f"خطأ في تنزيل تقرير الأعطال: {str(e)}", "danger")
        return redirect(url_for("camera_faults_sidebar"))

@app.route("/add_fault", methods=["POST"])
def add_fault():
    try:
        # الحصول على البيانات من النموذج
        camera_name = request.form.get('camera_name', '')
        problem = request.form.get('problem', '')
        
        if not camera_name or not problem:
            flash('اسم الكاميرا والمشكلة مطلوبين', 'danger')
            return redirect(request.referrer or url_for('home'))
        
        # إنشاء عطل جديد
        fault = Fault(
            camera_name=camera_name,
            problem=problem,
            description=f"عطل في كاميرا: {camera_name} - المشكلة: {problem}",
            fault_type='أخرى',
            device_type='كاميرا',
            reported_by='System'
        )
        
        db.session.add(fault)
        db.session.commit()
        
        flash(f'تم إضافة العطل #{fault.id} بنجاح', 'success')
        return redirect(request.referrer or url_for('home'))
        
    except Exception as e:
        flash(f'خطأ في إضافة العطل: {str(e)}', 'danger')
        return redirect(request.referrer or url_for('home'))

if __name__ == "__main__":
    try:
        app.run(debug=True, host='0.0.0.0', port=5001)
    except Exception as e:
        print(f"Failed to start server: {str(e)}")